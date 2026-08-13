# -*- coding: utf-8 -*-
"""
realestate_core.py - 전국 아파트 실거래가 조회 핵심 로직
(main.py의 Tkinter 데스크톱 앱을 웹앱용 로직으로 포팅)

app.py에서 이 모듈의 다음 함수/값을 가져다 씁니다.
    REGION_CODES
    fetch_transactions(service_key, lawd_cd, months, dong_filter, progress_cb) -> pd.DataFrame
    build_excel_bytes(df, sido, sigungu, dong, months) -> bytes
"""
import io
from datetime import datetime
from typing import Callable, Optional

import requests
import xml.etree.ElementTree as ET
import pandas as pd
from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

ProgressCB = Optional[Callable[[str], None]]

# =========================================================
# 전국 시·군·구 코드 (국토부 실거래가 API용 LAWD_CD)
# =========================================================
REGION_CODES = {
    "서울특별시": {
        "종로구": "11110", "중구": "11140", "용산구": "11170", "성동구": "11200",
        "광진구": "11215", "동대문구": "11230", "중랑구": "11260", "성북구": "11290",
        "강북구": "11305", "도봉구": "11320", "노원구": "11350", "은평구": "11380",
        "서대문구": "11410", "마포구": "11440", "양천구": "11470", "강서구": "11500",
        "구로구": "11530", "금천구": "11545", "영등포구": "11560", "동작구": "11590",
        "관악구": "11620", "서초구": "11650", "강남구": "11680", "송파구": "11710",
        "강동구": "11740",
    },
    "부산광역시": {
        "중구": "26110", "서구": "26140", "동구": "26170", "영도구": "26200",
        "부산진구": "26230", "동래구": "26260", "남구": "26290", "북구": "26320",
        "해운대구": "26350", "사하구": "26380", "금정구": "26410", "강서구": "26440",
        "연제구": "26470", "수영구": "26500", "사상구": "26530", "기장군": "26710",
    },
    "대구광역시": {
        "중구": "27110", "동구": "27140", "서구": "27170", "남구": "27200",
        "북구": "27230", "수성구": "27260", "달서구": "27290", "달성군": "27710",
        "군위군": "27720",
    },
    "인천광역시": {
        "중구": "28110", "동구": "28140", "미추홀구": "28177", "연수구": "28185",
        "남동구": "28200", "부평구": "28237", "계양구": "28245", "서구": "28260",
        "강화군": "28710", "옹진군": "28720",
    },
    "광주광역시": {
        "동구": "29110", "서구": "29140", "남구": "29155", "북구": "29170", "광산구": "29200",
    },
    "대전광역시": {
        "동구": "30110", "중구": "30140", "서구": "30170", "유성구": "30200", "대덕구": "30230",
    },
    "울산광역시": {
        "중구": "31110", "남구": "31140", "동구": "31170", "북구": "31200", "울주군": "31710",
    },
    "세종특별자치시": {"세종시": "36110"},
    "경기도": {
        "수원시 장안구": "41111", "수원시 권선구": "41113", "수원시 팔달구": "41115",
        "수원시 영통구": "41117", "성남시 수정구": "41131", "성남시 중원구": "41133",
        "성남시 분당구": "41135", "의정부시": "41150", "안양시 만안구": "41171",
        "안양시 동안구": "41173", "부천시": "41190", "광명시": "41210", "평택시": "41220",
        "동두천시": "41250", "안산시 상록구": "41271", "안산시 단원구": "41273",
        "고양시 덕양구": "41281", "고양시 일산동구": "41285", "고양시 일산서구": "41287",
        "과천시": "41290", "구리시": "41310", "남양주시": "41360", "오산시": "41370",
        "시흥시": "41390", "군포시": "41410", "의왕시": "41430", "하남시": "41450",
        "용인시 처인구": "41461", "용인시 기흥구": "41463", "용인시 수지구": "41465",
        "파주시": "41480", "이천시": "41500", "안성시": "41550", "김포시": "41570",
        "화성시": "41590", "광주시": "41610", "양주시": "41630", "포천시": "41650",
        "여주시": "41670", "연천군": "41800", "가평군": "41820", "양평군": "41830",
    },
    "강원특별자치도": {
        "춘천시": "51110", "원주시": "51130", "강릉시": "51150", "동해시": "51170",
        "태백시": "51190", "속초시": "51210", "삼척시": "51230", "홍천군": "51720",
        "횡성군": "51730", "영월군": "51750", "평창군": "51760", "정선군": "51770",
        "철원군": "51780", "화천군": "51790", "양구군": "51800", "인제군": "51810",
        "고성군": "51820", "양양군": "51830",
    },
    "충청북도": {
        "청주시 상당구": "43111", "청주시 서원구": "43112", "청주시 흥덕구": "43113",
        "청주시 청원구": "43114", "충주시": "43130", "제천시": "43150", "보은군": "43720",
        "옥천군": "43730", "영동군": "43740", "증평군": "43745", "진천군": "43750",
        "괴산군": "43760", "음성군": "43770", "단양군": "43800",
    },
    "충청남도": {
        "천안시 동남구": "44131", "천안시 서북구": "44133", "공주시": "44150",
        "보령시": "44180", "아산시": "44200", "서산시": "44210", "논산시": "44230",
        "계룡시": "44250", "당진시": "44270", "금산군": "44710", "부여군": "44760",
        "서천군": "44770", "청양군": "44790", "홍성군": "44800", "예산군": "44810",
        "태안군": "44825",
    },
    "전북특별자치도": {
        "전주시 완산구": "52111", "전주시 덕진구": "52113", "군산시": "52130",
        "익산시": "52140", "정읍시": "52180", "남원시": "52190", "김제시": "52210",
        "완주군": "52710", "진안군": "52720", "무주군": "52730", "장수군": "52740",
        "임실군": "52750", "순창군": "52770", "고창군": "52790", "부안군": "52800",
    },
    "전라남도": {
        "목포시": "46110", "여수시": "46130", "순천시": "46150", "나주시": "46170",
        "광양시": "46230", "담양군": "46710", "곡성군": "46720", "구례군": "46730",
        "고흥군": "46770", "보성군": "46780", "화순군": "46790", "장흥군": "46800",
        "강진군": "46810", "해남군": "46820", "영암군": "46830", "무안군": "46840",
        "함평군": "46860", "영광군": "46870", "장성군": "46880", "완도군": "46890",
        "진도군": "46900", "신안군": "46910",
    },
    "경상북도": {
        "포항시 남구": "47111", "포항시 북구": "47113", "경주시": "47130", "김천시": "47150",
        "안동시": "47170", "구미시": "47190", "영주시": "47210", "영천시": "47230",
        "상주시": "47250", "문경시": "47280", "경산시": "47290", "의성군": "47730",
        "청송군": "47750", "영양군": "47760", "영덕군": "47770", "청도군": "47820",
        "고령군": "47830", "성주군": "47840", "칠곡군": "47850", "예천군": "47900",
        "봉화군": "47920", "울진군": "47930", "울릉군": "47940",
    },
    "경상남도": {
        "창원시 의창구": "48121", "창원시 성산구": "48123", "창원시 마산합포구": "48125",
        "창원시 마산회원구": "48127", "창원시 진해구": "48129", "진주시": "48170",
        "통영시": "48220", "사천시": "48240", "김해시": "48250", "밀양시": "48270",
        "거제시": "48310", "양산시": "48330", "의령군": "48720", "함안군": "48730",
        "창녕군": "48740", "고성군": "48820", "남해군": "48840", "하동군": "48850",
        "산청군": "48860", "함양군": "48870", "거창군": "48880", "합천군": "48890",
    },
    "제주특별자치도": {"제주시": "50110", "서귀포시": "50130"},
}


# =========================================================
# 국토부 API 조회
# =========================================================

def _get_month_data(service_key: str, lawd_cd: str, deal_ymd: str):
    url = (
        "https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
        f"?serviceKey={service_key}"
        f"&LAWD_CD={lawd_cd}"
        f"&DEAL_YMD={deal_ymd}"
        "&numOfRows=9999"
        "&pageNo=1"
    )
    response = requests.get(url, timeout=30)
    if response.status_code != 200:
        raise RuntimeError(f"API 오류 (상태코드 {response.status_code})")
    xml_root = ET.fromstring(response.content)

    header = xml_root.find(".//header")
    if header is not None:
        result_code = header.findtext("resultCode", "")
        if result_code not in ("00", "000"):
            result_msg = header.findtext("resultMsg", "알 수 없는 오류")
            raise RuntimeError(f"API 오류: {result_msg} (코드 {result_code})")

    return xml_root.findall(".//item")


def fetch_transactions(
    service_key: str,
    lawd_cd: str,
    months: int,
    dong_filter: str = "",
    progress_cb: ProgressCB = None,
) -> pd.DataFrame:
    """지정 지역의 최근 N개월 아파트 매매 실거래가를 조회해 DataFrame으로 반환."""
    if not service_key:
        raise RuntimeError("실거래가 API 인증키(SERVICE_KEY)가 설정되지 않았습니다.")

    all_data = []
    today = datetime.today()
    dong_filter = (dong_filter or "").strip()

    for i in range(months):
        target_date = today - relativedelta(months=i)
        deal_ymd = target_date.strftime("%Y%m")
        if progress_cb:
            progress_cb(f"{deal_ymd} 실거래가 조회 중...")

        items = _get_month_data(service_key, lawd_cd, deal_ymd)

        for item in items:
            apt_name = (item.findtext("aptNm", "") or "").strip()
            dong_name = (item.findtext("umdNm", "") or "").strip()
            jibun = (item.findtext("jibun", "") or "").strip()
            deal_amount = item.findtext("dealAmount", "0") or "0"
            area = item.findtext("excluUseAr", "0") or "0"
            floor = item.findtext("floor", "") or ""
            year = item.findtext("dealYear", "") or ""
            month = item.findtext("dealMonth", "") or ""
            day = item.findtext("dealDay", "") or ""

            if dong_filter and dong_filter not in dong_name:
                continue

            try:
                deal_amount_manwon = int(deal_amount.replace(",", "").strip())
                area_m2 = float(area)
            except (TypeError, ValueError):
                continue

            deal_amount_won = deal_amount_manwon * 10000
            area_py = area_m2 / 3.3058
            price_per_py = deal_amount_won / area_py if area_py > 0 else 0

            try:
                contract_date = f"{year}-{int(month):02d}-{int(day):02d}"
            except (TypeError, ValueError):
                contract_date = ""

            all_data.append({
                "법정동": dong_name,
                "지번": jibun,
                "단지명": apt_name,
                "계약일": contract_date,
                "거래금액(원)": deal_amount_won,
                "전용면적(㎡)": round(area_m2, 2),
                "전용면적(평)": round(area_py, 2),
                "평당가(원)": round(price_per_py),
                "층": floor,
            })

    if not all_data:
        return pd.DataFrame()

    df = pd.DataFrame(all_data)
    df = df.sort_values(by="계약일", ascending=False).reset_index(drop=True)
    return df


# =========================================================
# 엑셀 보고서 생성 (main.py의 서식을 그대로 재현, 파일 대신 bytes로 반환)
# =========================================================

def build_excel_bytes(df: pd.DataFrame, sido: str, sigungu: str, dong: str, months) -> bytes:
    if df.empty:
        raise RuntimeError("내보낼 데이터가 없습니다. 먼저 조회를 실행해주세요.")

    export_df = df[[
        "법정동", "지번", "단지명", "계약일", "거래금액(원)",
        "전용면적(㎡)", "전용면적(평)", "평당가(원)", "층",
    ]].copy()
    export_df["비고"] = ""

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="실거래가 현황", index=False, startrow=9)
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb["실거래가 현황"]

    navy = "1F4E78"
    blue = "D9EAF7"
    light_blue = "EDF4F8"
    gray = "F2F2F2"
    white = "FFFFFF"
    dark_gray = "555555"
    border_color = "B7C3CC"
    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:J1")
    ws["A1"] = "아파트 실거래가 조사 결과"
    ws["A1"].font = Font(name="맑은 고딕", size=20, bold=True, color=white)
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    location_text = f"{sido} {sigungu}"
    if dong:
        location_text += f" {dong}"

    ws["A3"] = "조사지역"
    ws.merge_cells("B3:C3")
    ws["B3"] = location_text
    ws["D3"] = "조회기간"
    ws.merge_cells("E3:F3")
    ws["E3"] = f"최근 {months}개월"
    ws["G3"] = "조회일"
    ws.merge_cells("H3:J3")
    ws["H3"] = datetime.today().strftime("%Y-%m-%d")

    for cell_name in ["A3", "D3", "G3"]:
        cell = ws[cell_name]
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color=navy)
        cell.fill = PatternFill("solid", fgColor=gray)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for range_name in ["B3:C3", "E3:F3", "H3:J3"]:
        for row_cells in ws[range_name]:
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24

    count = len(df)
    avg_price = int(df["거래금액(원)"].mean())
    avg_py_price = int(df["평당가(원)"].mean())
    max_price = int(df["거래금액(원)"].max())
    min_price = int(df["거래금액(원)"].min())

    kpi_ranges = [
        ("A5:B5", "A6:B6", "총 거래건수", count),
        ("C5:D5", "C6:D6", "평균 거래가", avg_price),
        ("E5:F5", "E6:F6", "평균 평당가", avg_py_price),
        ("G5:H5", "G6:H6", "최고 거래가", max_price),
        ("I5:J5", "I6:J6", "최저 거래가", min_price),
    ]
    for title_range, value_range, title_text, value in kpi_ranges:
        ws.merge_cells(title_range)
        ws.merge_cells(value_range)
        title_cell = ws[title_range.split(":")[0]]
        value_cell = ws[value_range.split(":")[0]]
        title_cell.value = title_text
        value_cell.value = value
        title_cell.font = Font(name="맑은 고딕", size=10, bold=True, color=navy)
        title_cell.fill = PatternFill("solid", fgColor=blue)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.font = Font(name="맑은 고딕", size=12, bold=True, color="222222")
        value_cell.fill = PatternFill("solid", fgColor=white)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_cells in ws[title_range]:
            for cell in row_cells:
                cell.border = border
        for row_cells in ws[value_range]:
            for cell in row_cells:
                cell.border = border

    ws["A6"].number_format = '#,##0"건"'
    for cell_name in ["C6", "E6", "G6", "I6"]:
        ws[cell_name].number_format = '#,##0"원"'
    ws.row_dimensions[5].height = 23
    ws.row_dimensions[6].height = 30

    ws.merge_cells("A8:J8")
    ws["A8"] = "※ 국토교통부 실거래가 공개자료 기준 | 금액은 원 단위 | 비고란은 사용자 입력용"
    ws["A8"].font = Font(name="맑은 고딕", size=9, color=dark_gray)
    ws["A8"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[8].height = 20

    header_row = 10
    for cell in ws[header_row]:
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 27

    max_row = ws.max_row
    for row in range(header_row + 1, max_row + 1):
        row_fill = PatternFill("solid", fgColor=light_blue if row % 2 == 0 else white)
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.fill = row_fill
            cell.border = border
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="center")
        for col in [1, 2, 4, 6, 7, 9]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=5).number_format = "#,##0"
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=6).number_format = "0.00"
        ws.cell(row=row, column=7).number_format = "0.00"
        ws.cell(row=row, column=8).number_format = "#,##0"
        ws.cell(row=row, column=8).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=10).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 21

    widths = {"A": 14, "B": 13, "C": 32, "D": 14, "E": 22, "F": 16, "G": 16, "H": 22, "I": 9, "J": 25}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.auto_filter.ref = f"A{header_row}:J{max_row}"
    ws.freeze_panes = "A11"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "10:10"

    # ---- 단지별 요약 시트 ----
    summary = wb.create_sheet("단지별 요약")
    summary.merge_cells("A1:G1")
    summary["A1"] = "단지별 실거래가 요약"
    summary["A1"].font = Font(name="맑은 고딕", size=18, bold=True, color=white)
    summary["A1"].fill = PatternFill("solid", fgColor=navy)
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 35

    summary.merge_cells("A2:G2")
    summary["A2"] = f"조사지역 : {location_text}   |   조회기간 : 최근 {months}개월"
    summary["A2"].font = Font(name="맑은 고딕", size=10, color=dark_gray)
    summary["A2"].alignment = Alignment(horizontal="center", vertical="center")

    apartment_summary = (
        df.groupby(["법정동", "단지명"])
        .agg(
            거래건수=("거래금액(원)", "count"),
            평균거래가=("거래금액(원)", "mean"),
            평균평당가=("평당가(원)", "mean"),
            최고거래가=("거래금액(원)", "max"),
            최저거래가=("거래금액(원)", "min"),
        )
        .reset_index()
    )
    apartment_summary = apartment_summary.sort_values(by=["거래건수", "평균거래가"], ascending=[False, False])

    headers = ["법정동", "단지명", "거래건수", "평균거래가(원)", "평균평당가(원)", "최고거래가(원)", "최저거래가(원)"]
    summary.append([])
    summary.append(headers)
    for _, row in apartment_summary.iterrows():
        summary.append([
            row["법정동"], row["단지명"], int(row["거래건수"]),
            int(row["평균거래가"]), int(row["평균평당가"]),
            int(row["최고거래가"]), int(row["최저거래가"]),
        ])

    summary_header_row = 4
    for cell in summary[summary_header_row]:
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[summary_header_row].height = 27

    for row in range(5, summary.max_row + 1):
        fill = PatternFill("solid", fgColor=light_blue if row % 2 == 0 else white)
        for col in range(1, 8):
            cell = summary.cell(row=row, column=col)
            cell.border = border
            cell.fill = fill
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="center")
        summary.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        summary.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        summary.cell(row=row, column=3).number_format = '#,##0"건"'
        for col in [4, 5, 6, 7]:
            summary.cell(row=row, column=col).number_format = "#,##0"
            summary.cell(row=row, column=col).alignment = Alignment(horizontal="right", vertical="center")
        summary.row_dimensions[row].height = 21

    summary_widths = {"A": 15, "B": 35, "C": 13, "D": 22, "E": 22, "F": 22, "G": 22}
    for col, width in summary_widths.items():
        summary.column_dimensions[col].width = width

    summary.auto_filter.ref = f"A4:G{summary.max_row}"
    summary.freeze_panes = "A5"
    summary.sheet_view.showGridLines = False
    summary.sheet_view.zoomScale = 90

    out_buf = io.BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue()
