# -*- coding: utf-8 -*-
"""
collector_core.py - 정비사업 입찰공고 수집기 핵심 로직 (v3 로직 반영판)
--------------------------------------------------------------------------
app.py(Streamlit 화면)에서 다음 함수/값들을 가져다 씁니다.

    load_config() / save_config(cfg)
    run_pipeline(service_key, cfg, yesterday_only, days_back, download_attachments, progress_cb)
    search_by_institution(institution_keyword, service_key, cfg, years_back, biz_types,
                           download_attachments, progress_cb)
    estimate_institution_search_calls(years_back, biz_types)
    DEFAULT_ATTACHMENT_DIR, LOG_DIR

[이 버전(v3)에서 반영된 사용자 지정 로직]
  - 공고구분 필드로 재공고/정정공고를 정확히 판정 (변경·취소·철회·무효만 제외)
  - 취소공고가 뒤늦게 올라오면, 이미 수집된 원공고를 찾아 무효화
  - "OO공사 용역"처럼 실제로는 공사 발주인 건은 제외 (감리/설계/심의 등 딸린 용역은 유지)
  - 설계 세부내역을 공고명에 등장하는 분야명으로 자동 조합 (예: "전기+통신")
  - 협력업체 공고 중 구분이 "기타"로 남은 건을, 다운로드한 공고문 본문의 용역명으로 재분류
  - 공고문/지침서 본문에서 지역·면적(구역/대지/건축연면적) 자동 추출

[네트워크]
  requests로 우선 요청하고, 실패하면(사내망 등에서 https가 막히는 경우) curl로 재시도합니다.
  (기존에 "회사 PC에서는 curl은 되는데 requests는 안 된다"는 사례가 있어 안전장치로 남겨둠)
"""

import os
import re
import time
import html
import json
import shutil
import subprocess
import urllib.parse
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, Optional

import requests
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================== 경로 ===============================
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
LOG_DIR = BASE_DIR / "logs"
CONFIG_PATH = BASE_DIR / "config.json"
DEFAULT_ATTACHMENT_DIR = OUTPUT_DIR / "공고문_다운로드"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(parents=True, exist_ok=True)

ProgressCB = Optional[Callable[[str], None]]


# ============================== 로그 ===============================
class _Logger:
    """진행 메시지를 로그 파일에 남기고, 있으면 progress_cb(웹 화면 실시간 표시)도 호출."""

    def __init__(self, prefix: str, progress_cb: ProgressCB):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.path = LOG_DIR / f"collector_{prefix}_{ts}.log"
        self._cb = progress_cb
        try:
            self._fh = open(self.path, "w", encoding="utf-8")
        except OSError:
            self._fh = None

    def log(self, msg: str):
        line = f"[{datetime.now():%H:%M:%S}] {msg}"
        if self._fh:
            try:
                self._fh.write(line + "\n")
                self._fh.flush()
            except OSError:
                pass
        if self._cb:
            try:
                self._cb(msg)
            except Exception:
                pass

    def close(self):
        if self._fh:
            try:
                self._fh.close()
            except OSError:
                pass


# ============================== 네트워크 ===============================
# requests를 우선 쓰고, 실패하면(사내망 TLS 차단 등) curl.exe / curl로 재시도합니다.
CURL_PATH = "curl"


def _build_full_url(url: str, params: dict) -> str:
    """
    공공데이터포털 "일반 인증키(Encoding)"는 이미 URL 인코딩된 값(%2B, %2F, %3D 등 포함)입니다.
    urlencode()로 한 번 더 인코딩하면 %가 %25로 다시 인코딩되어 키가 깨지므로,
    serviceKey만 인코딩 없이 그대로 붙이고 나머지 파라미터만 urlencode로 인코딩합니다.
    """
    params = dict(params)
    service_key = params.pop("serviceKey", "")
    query = urllib.parse.urlencode(params, safe="")
    if service_key:
        return f"{url}?serviceKey={service_key}&{query}" if query else f"{url}?serviceKey={service_key}"
    return f"{url}?{query}"


def _http_get_json(url: str, params: dict, timeout: int = 30):
    """(raw_text, parsed_json_or_None)을 반환. 완전히 실패하면 RuntimeError."""
    full_url = _build_full_url(url, params)

    try:
        resp = requests.get(full_url, timeout=timeout)
        resp.raise_for_status()
        return resp.text, _try_json(resp.text)
    except Exception as e:
        if not shutil.which(CURL_PATH):
            raise RuntimeError(f"요청 실패: {e}")
        try:
            result = subprocess.run(
                [CURL_PATH, "-s", "-S", "--max-time", str(timeout), full_url],
                capture_output=True,
                timeout=timeout + 10,
            )
        except (FileNotFoundError, subprocess.TimeoutExpired) as ce:
            raise RuntimeError(f"requests 실패({e}) / curl도 실패({ce})")
        if result.returncode != 0:
            err_text = result.stderr.decode("utf-8", errors="ignore").strip()
            raise RuntimeError(f"requests 실패({e}) / curl 오류(code={result.returncode}): {err_text[:200]}")
        raw_text = result.stdout.decode("utf-8", errors="ignore")
        return raw_text, _try_json(raw_text)


def _try_json(raw_text: str):
    try:
        return json.loads(raw_text)
    except ValueError:
        return None


def _download_file(url: str, dest_path: str, timeout: int = 30) -> bool:
    """requests로 우선 다운로드하고, 실패하면 curl로 재시도. 성공 여부(bool) 반환."""
    try:
        resp = requests.get(url, timeout=timeout, stream=True)
        resp.raise_for_status()
        with open(dest_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=65536):
                if chunk:
                    f.write(chunk)
        return os.path.exists(dest_path) and os.path.getsize(dest_path) > 0
    except Exception:
        if not shutil.which(CURL_PATH):
            return False
        try:
            result = subprocess.run(
                [CURL_PATH, "-s", "-S", "-L", "--max-time", str(timeout), "-o", str(dest_path), url],
                capture_output=True,
                timeout=timeout + 10,
            )
        except (FileNotFoundError, subprocess.TimeoutExpired):
            return False
        if result.returncode != 0:
            return False
        return os.path.exists(dest_path) and os.path.getsize(dest_path) > 0


# ============================ 분류 기준 (코드 상수 - 관리자 편집 대상 아님) ============================

# 공고 상태 판정
NOTICE_KIND_FIELDS = [
    "ntceKindNm", "bidNtceKindNm", "pblancKindNm", "ntceDivNm",
    "bidNtceDivNm", "rgstTyNm", "ntceSeCd", "ntceSeNm",
]
EXCLUDE_NOTICE_KINDS = ["변경", "취소", "철회", "무효"]
FALLBACK_EXCLUDE_STATUS_KEYWORDS = ["변경공고", "변경 공고", "취소공고", "취소 공고"]
CANCEL_MARKS = ["취소", "철회", "무효"]
CHANGE_MARKS = ["변경"]

# 공사 발주 건 제외 규칙
CONSTRUCTION_HINT = "공사"
CONSTRUCTION_ALLOW = [
    "감리", "설계", "계획서", "심의", "평가", "진단", "점검",
    "용역관리", "사업관리", "건설사업관리",
]
EXCLUDE_UNLESS_KEYWORD = {"경관": "경관심의"}
EXCLUDE_TITLE_PHRASES = ["임대주택 매각"]
UNCLEAR_COOP_NOTE = "수동확인"

# 설계 세부내역 자동 조합
DESIGN_FIELDS = [
    ("정보통신", "통신"), ("통신", "통신"), ("전기", "전기"), ("소방", "소방"),
    ("기계", "기계설비"), ("설비", "기계설비"),
    ("토목", "토목"), ("조경", "조경"), ("건축", "건축"),
]

ENGINEERING_KEYWORDS = [
    "교통영향평가", "환경영향평가", "재해영향평가", "교육환경평가", "교육영향평가",
    "친환경평가", "친환경인증", "경관심의", "지하안전", "전력계통영향평가", "전력계통",
    "건축물안전영향영가", "건축물안전영향평가", "해체계획서", "해체계획",
]

# ============================ 필터 키워드 3종 - config.json에서 관리, 아래는 최초 기본값 ============================

DEFAULT_INCLUDE_INSTITUTION_KEYWORDS = [
    "재개발", "재건축", "정비사업", "도심복합", "주택정비", "소규모재건축", "가로주택", "소규모주택",
    "자산신탁", "토지신탁", "부동산신탁", "코리아신탁", "대한토지신탁", "하나자산신탁",
    "케이비부동산신탁", "한국자산신탁",
    "추진위원회", "준비위원회", "추진준비위", "주민대표회의", "운영위원회",
]

DEFAULT_INCLUDE_TITLE_KEYWORDS = [
    "CM", "PCM", "건설사업관리",
    "PM", "P.M", "사업관리",
    "설계",
    "설계자", "설계업자", "건축설계", "기본설계", "실시설계",
    "감리",
    "환경영향평가", "재해영향평가", "교육영향평가", "교육환경평가",
    "교통영향평가", "지하안전", "건축물안전영향평가", "전력계통영향평가",
    "친환경",
    "토목설계", "토목",
    "흙막이", "석면해체", "해체계획", "인허가", "경관심의", "기반시설",
    "소방", "전기", "통신",
]

DEFAULT_EXCLUDE_TITLE_KEYWORDS = [
    "정비사업전문관리", "시공사", "시공자",
    "세무", "회계", "법", "변호사", "소송", "매도청구",
    "이주관리", "범죄예방", "토지등소유자", "전체회의", "총회", "홍보", "채용",
    "분양", "책임매입", "매각", "주택관리업자", "사업시행자", "우선협상",
    "정비계획", "CCTV", "횡단보도", "감정평가", "조합설립", "청산",
    "보류지", "상가", "보험", "폐기물", "공사비", "국공유지", "무상양도", "기부대양여",
    "공공디자인", "도시계획",
    "BF인증", "장애물 없는", "설계공모", "현상설계",
    "석면사전조사", "석면측정", "농도측정", "농도 측정", "HUG보증", "지장물",
    "내진설계", "내풍설계", "풍동실험", "지적측량", "토질",
    "관리처분", "관리계획수립", "원인자부담금", "음식물",
    "인테리어", "실내건축", "가구설계",
    "통신보안", "보안솔루션", "보안 솔루션", "정보보안", "홈네트워크 보안",
    "사업비절감", "사업비 절감", "정산", "환급", "원가절감",
    "도료", "도장", "방수", "조명", "사인물",
    "제연", "부속실",
    "토양오염", "토양정화", "정화검증", "오염토", "오염토양",
    "환경오염", "수질오염", "지하수정화", "오염도조사",
]


def load_config() -> dict:
    """config.json을 읽어옵니다. 없으면 기본 키워드로 새로 만듭니다."""
    cfg = {}
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
        except (OSError, ValueError):
            cfg = {}
    changed = False
    if "include_institution_keywords" not in cfg:
        cfg["include_institution_keywords"] = list(DEFAULT_INCLUDE_INSTITUTION_KEYWORDS)
        changed = True
    if "include_title_keywords" not in cfg:
        cfg["include_title_keywords"] = list(DEFAULT_INCLUDE_TITLE_KEYWORDS)
        changed = True
    if "exclude_title_keywords" not in cfg:
        cfg["exclude_title_keywords"] = list(DEFAULT_EXCLUDE_TITLE_KEYWORDS)
        changed = True
    if changed:
        try:
            save_config(cfg)
        except OSError:
            pass
    return cfg


def save_config(cfg: dict):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# ============================ 조합별 이력조회 - 즐겨찾기/최근조회기록 ============================

HISTORY_STORE_PATH = BASE_DIR / "hist_history.json"
MAX_RECENT_SEARCHES = 15


def _load_hist_store() -> dict:
    if HISTORY_STORE_PATH.exists():
        try:
            with open(HISTORY_STORE_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, ValueError):
            data = {}
    else:
        data = {}
    data.setdefault("favorites", [])
    data.setdefault("recent", [])
    return data


def _save_hist_store(data: dict):
    with open(HISTORY_STORE_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_favorites() -> list:
    return _load_hist_store()["favorites"]


def add_favorite(name: str):
    name = name.strip()
    if not name:
        return
    data = _load_hist_store()
    if name not in data["favorites"]:
        data["favorites"].append(name)
        _save_hist_store(data)


def remove_favorite(name: str):
    data = _load_hist_store()
    if name in data["favorites"]:
        data["favorites"].remove(name)
        _save_hist_store(data)


def get_recent_searches() -> list:
    return _load_hist_store()["recent"]


def add_recent_search(name: str):
    name = name.strip()
    if not name:
        return
    data = _load_hist_store()
    recent = [r for r in data["recent"] if r != name]
    recent.insert(0, name)
    data["recent"] = recent[:MAX_RECENT_SEARCHES]
    _save_hist_store(data)


# ============================ API 호출 설정 ============================
URL_PREFIX_CANDIDATES_DEFAULT = [
    "https://apis.data.go.kr/1230000/ao/PrvtBidNtceService",
    "http://apis.data.go.kr/1230000/ao/PrvtBidNtceService",
]
OPERATION_BY_BIZ = {
    "용역": "getPrvtBidPblancListInfoServc",
    "물품": "getPrvtBidPblancListInfoThng",
    "공사": "getPrvtBidPblancListInfoCnstwk",
    "기타": "getPrvtBidPblancListInfoEtc",
}
CHUNK_DAYS = 3
NUM_OF_ROWS = 500

FIELD_CANDIDATES = {
    "공고명": ["bidNtceNm", "pblancNm", "ntceNm"],
    "공고구분": NOTICE_KIND_FIELDS,
    "발주기관": ["ntceInsttNm", "dminsttNm", "ordersInsttNm", "orderInsttNm"],
    "지역": ["prtcptPsblRgnNm", "rgnNm", "areaNm"],
    "공고일시": ["bidNtceDt", "nticeDt", "ntceDt", "pblancDt", "rgstDt"],
    "추정가격": ["asignBdgtAmt", "presmptPrce"],
    "기초금액": ["bssamt", "baseAmt", "presmptPrce"],
    "투찰마감": ["bidClseDt", "bidClseDate", "opengDt"],
    "공고번호": ["bidNtceNo", "pblancNo"],
    "공고차수": ["bidNtceOrd", "pblancOrd"],
}


def pick_field(item: dict, candidates: list, default=""):
    for c in candidates:
        if c in item and item[c] not in (None, ""):
            return item[c]
    return default


def clean_text(s) -> str:
    if not s:
        return ""
    s = html.unescape(str(s))
    s = s.replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ---------------------------- 날짜 계산 ----------------------------

def get_previous_business_day(reference_dt: datetime) -> datetime:
    day = reference_dt.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
    while day.weekday() >= 5:
        day -= timedelta(days=1)
    return day


def _make_date_chunks_for_yesterday(look_ahead_for_cancel: bool = True):
    target_day = get_previous_business_day(datetime.now())
    start_dt = target_day
    end_dt = target_day + timedelta(days=1, minutes=-1)
    if look_ahead_for_cancel:
        end_dt = max(end_dt, datetime.now())
    return target_day, [(start_dt, end_dt)]


def _make_date_chunks_for_range(start_dt: datetime, end_dt: datetime, chunk_days: int = CHUNK_DAYS):
    chunks = []
    cur = start_dt
    while cur < end_dt:
        chunk_end = min(cur + timedelta(days=chunk_days), end_dt)
        chunks.append((cur, chunk_end))
        cur = chunk_end
    return chunks


def estimate_institution_search_calls(years_back: float, biz_types: list) -> int:
    """이력조회 화면에 "예상 API 호출 횟수"를 보여주기 위한 대략적인 추정."""
    total_days = max(1, int(round(years_back * 365)))
    n_chunks = max(1, -(-total_days // CHUNK_DAYS))  # 올림 나눗셈
    return n_chunks * max(1, len(biz_types or ["용역"]))


# ---------------------------- 공고 상태 / 판정 ----------------------------

def get_notice_kind(item: dict) -> str:
    return clean_text(pick_field(item, NOTICE_KIND_FIELDS))


def notice_status(item: dict, title: str) -> str:
    kind = get_notice_kind(item)
    hay = kind if kind else title.replace(" ", "")
    if any(k in hay for k in CANCEL_MARKS):
        return "cancel"
    if kind:
        if any(k in kind for k in CHANGE_MARKS):
            return "change"
    elif "변경공고" in hay:
        return "change"
    return "normal"


def make_cancel_key(title: str, institution: str) -> str:
    t = re.sub(r"[\s\(\)\[\]{}·,\.\-_/]", "", str(title))
    tail = r"(재공고|입찰공고|입찰|공고|취소|철회|무효|변경|정정)$"
    while True:
        new = re.sub(tail, "", t)
        if new == t:
            break
        t = new
    inst = re.sub(r"\s", "", str(institution))
    return f"{t}|{inst}"


def make_content_key(title: str, institution: str) -> str:
    t = re.sub(r"[\s\(\)\[\]{}·,\.\-_/]", "", str(title))
    t = re.sub(r"(입찰재공고|입찰공고|재공고|입찰|공고)$", "", t)
    inst = re.sub(r"\s", "", str(institution))
    return f"{t}|{inst}"


def institution_passes(institution: str, include_institution_keywords: list) -> bool:
    return any(kw in institution for kw in include_institution_keywords)


def title_excluded(title: str, exclude_title_keywords: list) -> bool:
    t = title.replace(" ", "")
    if any(kw.replace(" ", "") in t for kw in exclude_title_keywords):
        return True
    if any(phrase.replace(" ", "") in t for phrase in EXCLUDE_TITLE_PHRASES):
        return True
    for bad_kw, unless_kw in EXCLUDE_UNLESS_KEYWORD.items():
        if bad_kw in t and unless_kw not in t:
            return True
    if CONSTRUCTION_HINT in t and not any(a in t for a in CONSTRUCTION_ALLOW):
        return True
    return False


def title_included(title: str, include_title_keywords: list) -> bool:
    return any(kw in title for kw in include_title_keywords)


def classify_title(title: str):
    """(구분, 세부내역)을 결정. 매칭 안 되면 ("기타", "")."""
    t = re.sub(r"[\s,·]", "", title)

    if any(kw in t for kw in ["건설사업관리", "PCM", "CM"]):
        return "CM", "건설사업관리"
    if any(kw in t for kw in ["사업관리", "P.M", "PM"]):
        return "PM", "사업관리"

    if "감리" in t:
        return "감리", ("석면해체" if "석면" in t else "감리")

    if "설계" in t:
        if "설계자" in t or "설계업자" in t:
            return "설계", "설계자"
        found = {}
        for kw, name in DESIGN_FIELDS:
            if kw == "건축":
                # [버그 수정] "재건축"의 "건축"과 겹쳐 항상 오탐되는 문제 방지.
                #   "재건축"이 아닌 진짜 "건축"(예: "건축설계", "건축물") 등장 위치만 인정.
                m = re.search(r"(?<!재)건축", t)
                pos = m.start() if m else -1
            else:
                pos = t.find(kw)
            if pos >= 0 and (name not in found or pos < found[name]):
                found[name] = pos
        fields = [name for name, _ in sorted(found.items(), key=lambda x: x[1])]
        return "설계", ("+".join(fields) if fields else "설계")

    for kw in ENGINEERING_KEYWORDS:
        if kw in t:
            return "엔지니어링", kw

    return "기타", ""


def extract_region(institution: str, title: str) -> str:
    text = institution + " " + title
    m = re.search(r"[가-힣]+시\s?[가-힣]+구", text)
    if m:
        return normalize_region_name(m.group(0).replace("  ", " "))
    m = re.search(r"[가-힣]+\d*동\d*가?\s*\d+(-\d+)?번지", text)
    if m:
        return m.group(0)
    m = re.search(r"[가-힣0-9]+구역", text)
    if m:
        return m.group(0)
    m = re.search(r"[가-힣]+\d*동(?=\s|$)", text)
    if m:
        return m.group(0)
    return ""


def format_notice_datetime(raw) -> str:
    if not raw:
        return ""
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) >= 8:
        return f"{digits[0:4]}-{digits[4:6]}-{digits[6:8]}"
    return str(raw)


def extract_project_type(text: str) -> str:
    order = [
        "소규모재건축", "가로주택정비사업", "가로주택",
        "도심복합개발", "도심복합",
        "재건축", "재개발", "주택정비사업", "주택정비",
    ]
    for kw in order:
        if kw in text:
            return kw
    return ""


# ---------------------------- API 호출 ----------------------------

def fetch_one_page(operation: str, begin_dt: datetime, end_dt: datetime, page_no: int,
                    service_key: str, url_prefixes: list):
    params = {
        "serviceKey": service_key,
        "type": "json",
        "inqryDiv": "1",
        "inqryBgnDt": begin_dt.strftime("%Y%m%d%H%M"),
        "inqryEndDt": end_dt.strftime("%Y%m%d%H%M"),
        "pageNo": page_no,
        "numOfRows": NUM_OF_ROWS,
    }
    last_error = None
    for prefix in list(url_prefixes):
        url = f"{prefix}/{operation}"
        try:
            raw_text, data = _http_get_json(url, params, timeout=45)
        except RuntimeError as e:
            last_error = f"{e} (주소: {prefix})"
            continue
        if data is None:
            last_error = f"JSON 파싱 실패 (주소: {prefix}) - 응답 앞부분: {raw_text[:300]}"
            continue
        if "response" not in data:
            last_error = f"예상치 못한 응답 구조 (주소: {prefix})"
            continue
        header = data["response"].get("header", {})
        result_code = header.get("resultCode")
        if result_code != "00":
            last_error = f"API 오류 resultCode={result_code}, msg={header.get('resultMsg')} (주소: {prefix})"
            continue
        if prefix in url_prefixes:
            url_prefixes.remove(prefix)
            url_prefixes.insert(0, prefix)
        body = data.get("response", {}).get("body", {})
        items = body.get("items", [])
        if isinstance(items, dict):
            items = [items]
        total_count = int(body.get("totalCount", 0) or 0)
        return items, total_count
    raise RuntimeError(last_error or "알 수 없는 오류 (모든 주소 후보 실패)")


def fetch_all(biz_type: str, date_chunks: list, service_key: str, logger: "_Logger" = None):
    operation = OPERATION_BY_BIZ[biz_type]
    url_prefixes = list(URL_PREFIX_CANDIDATES_DEFAULT)
    all_items = []
    for begin_dt, end_dt in date_chunks:
        page_no = 1
        while True:
            try:
                items, total_count = fetch_one_page(operation, begin_dt, end_dt, page_no, service_key, url_prefixes)
            except RuntimeError as e:
                if logger:
                    logger.log(f"  [!] {biz_type} 조회 중 오류: {e}")
                break
            if not items:
                break
            all_items.extend(items)
            if page_no * NUM_OF_ROWS >= total_count:
                break
            page_no += 1
            time.sleep(0.15)
    for it in all_items:
        it["_업무구분"] = biz_type
    return all_items


def collect_raw(biz_types: list, date_chunks: list, service_key: str, logger: "_Logger" = None):
    all_rows = []
    for biz_type in biz_types:
        if logger:
            logger.log(f"조회 중: [{biz_type}] ...")
        rows = fetch_all(biz_type, date_chunks, service_key, logger=logger)
        if logger:
            logger.log(f"  -> {len(rows)}건 조회됨")
        all_rows.extend(rows)
    return all_rows


# ---------------------------- 필터링 / 분류 ----------------------------

def apply_filters_and_classify(
    raw_items: list,
    include_institution_keywords: list,
    include_title_keywords: list,
    exclude_title_keywords: list,
    institution_match_fn=None,
    restrict_to_date: Optional[str] = None,
    logger: "_Logger" = None,
) -> pd.DataFrame:
    """
    institution_match_fn(institution:str)->bool 을 넘기면 발주기관 판정에 그 함수를 씁니다.
    (이력조회에서는 "발주기관명에 검색어가 포함되는지"로 대체합니다)
    restrict_to_date='YYYY-MM-DD'를 넘기면 그 날짜 공고만 최종 남깁니다. (오늘의 공고 수집용)
    """
    if institution_match_fn is None:
        def institution_match_fn(inst):
            return institution_passes(inst, include_institution_keywords)

    records = []
    cancel_keys = {}

    for it in raw_items:
        title = clean_text(pick_field(it, FIELD_CANDIDATES["공고명"]))
        institution = clean_text(pick_field(it, FIELD_CANDIDATES["발주기관"]))
        if not title:
            continue

        status = notice_status(it, title)
        if status in ("cancel", "change"):
            cancel_keys[make_cancel_key(title, institution)] = (
                "취소공고" if status == "cancel" else "변경공고"
            )
            continue

        if not institution_match_fn(institution):
            continue
        if title_excluded(title, exclude_title_keywords):
            continue

        note = ""
        is_coop = "협력업체" in title
        has_clear_keyword = title_included(title, include_title_keywords)
        if is_coop:
            if not has_clear_keyword:
                note = UNCLEAR_COOP_NOTE
        else:
            if not has_clear_keyword:
                continue

        category, detail = classify_title(title)
        region = extract_region(institution, title)

        note_parts = []
        if note:
            note_parts.append(note)
        kind = get_notice_kind(it)
        for mark in ("재공고", "정정"):
            if mark in kind or mark in title.replace(" ", ""):
                note_parts.append(f"[{mark if mark != '정정' else '정정공고'}]")
                break
        if category and category != "기타":
            tag = f"[{category}/{detail}]" if detail else f"[{category}]"
            note_parts.append(tag)
        remark = " ".join(note_parts)

        notice_dt = format_notice_datetime(pick_field(it, FIELD_CANDIDATES["공고일시"]))
        close_dt = format_notice_datetime(pick_field(it, FIELD_CANDIDATES["투찰마감"]))
        folder_digits = re.sub(r"\D", "", notice_dt)[:8] if notice_dt else ""

        rec = {
            "폴더명": folder_digits,
            "구분": category,
            "세부내역": detail,
            "공고일시": notice_dt,
            "마감일시": close_dt,
            "현장설명회일": "",
            "지역": region,
            "공고명": title,
            "발주기관": institution,
            "업체명": "",
            "건축연면적(㎡)": "",
            "건축연면적(평)": "",
            "대지면적(㎡)": "",
            "대지면적(평)": "",
            "구역면적(㎡)": "",
            "구역면적(평)": "",
            "입찰금액(원)": "",
            "평단가(원)": "",
            "비고": remark,
            "_공고번호": pick_field(it, FIELD_CANDIDATES["공고번호"]),
            "_공고차수": pick_field(it, FIELD_CANDIDATES["공고차수"]),
            "_업무구분": it.get("_업무구분", ""),
            "_구분": category,
            "_세부내역": detail,
            "_사업유형": extract_project_type(title + " " + institution),
            "_raw_dict": it,
            "_취소키": make_cancel_key(title, institution),
        }
        records.append(rec)

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)

    if cancel_keys:
        hit = df["_취소키"].isin(cancel_keys.keys())
        if logger:
            for _, r in df[hit].iterrows():
                logger.log(f"  [{cancel_keys[r['_취소키']]}] 확인되어 제외: {r['공고명']} ({r['발주기관']})")
        df = df[~hit]
        if df.empty:
            return pd.DataFrame()

    if restrict_to_date:
        before = len(df)
        df = df[df["공고일시"] == restrict_to_date]
        if logger and before != len(df):
            logger.log(f"  [기간정리] 대상일({restrict_to_date}) 외 공고 {before - len(df)}건 제외")
        if df.empty:
            return pd.DataFrame()

    df["_고유키"] = (
        df["_공고번호"].astype(str) + "_" + df["_공고차수"].astype(str) + "_" + df["_업무구분"]
    )
    df = df.drop_duplicates(subset="_고유키", keep="first")

    df["_내용키"] = df.apply(lambda r: make_content_key(r["공고명"], r["발주기관"]), axis=1)
    before = len(df)
    df = df.drop_duplicates(subset="_내용키", keep="first")
    if logger and before != len(df):
        logger.log(f"  [중복정리] 내용이 같은 공고 {before - len(df)}건을 제외했습니다.")

    return df


# ---------------------------- 첨부파일(공고문/지침서) 다운로드 ----------------------------

def sanitize_filename(name: str, max_len: int = 80) -> str:
    name = clean_text(name)
    name = re.sub(r'[\\/:*?"<>|]', "", name)
    name = name.strip(" .")
    if len(name) > max_len:
        name = name[:max_len]
    return name or "제목없음"


def find_urls_in_item(item: dict):
    found = []
    for key, value in item.items():
        if key.startswith("_"):
            continue
        if isinstance(value, str) and value.strip().lower().startswith("http"):
            found.append((key, value.strip()))
    return found


def guess_extension(url: str, content_type: str, content_bytes: bytes = b"") -> str:
    if content_bytes:
        head = content_bytes[:8]
        if head.startswith(b"%PDF"):
            return ".pdf"
        if head.startswith(b"PK\x03\x04"):
            return ".hwpx"
        if head.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
            return ".hwp"
    for ext in (".pdf", ".hwp", ".hwpx", ".zip", ".doc", ".docx", ".xls", ".xlsx"):
        if url.lower().endswith(ext):
            return ext
    if content_type:
        ct = content_type.lower()
        if "pdf" in ct:
            return ".pdf"
        if "zip" in ct:
            return ".zip"
        if "hwp" in ct:
            return ".hwp"
    return ".dat"


def extract_text_from_pdf(file_path: str) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            return ""
    try:
        reader = PdfReader(file_path)
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception:
        return ""


def extract_text_from_hwp(file_path: str) -> str:
    try:
        import olefile
    except ImportError:
        return ""
    try:
        ole = olefile.OleFileIO(file_path)
        stream_name = None
        for name in ole.listdir():
            if "/".join(name).lower().endswith("prvtext"):
                stream_name = name
                break
        if not stream_name:
            return ""
        data = ole.openstream(stream_name).read()
        return data.decode("utf-16le", errors="ignore")
    except Exception:
        return ""


def extract_text_from_hwpx(file_path: str) -> str:
    try:
        import zipfile
        chunks = []
        with zipfile.ZipFile(file_path) as z:
            for name in z.namelist():
                if name.startswith("Contents/section") and name.endswith(".xml"):
                    with z.open(name) as f:
                        chunks.append(f.read().decode("utf-8", errors="ignore"))
        return "\n".join(chunks)
    except Exception:
        return ""


def extract_text_from_file(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        return extract_text_from_pdf(file_path)
    if ext == ".hwp":
        return extract_text_from_hwp(file_path)
    if ext == ".hwpx":
        return extract_text_from_hwpx(file_path)
    return ""


# ---- 공고문 본문에서 "용역명" 뽑아 협력업체 공고 분류하기 ----

SERVICE_NAME_LABELS = ["용역명", "용역명칭", "과업명", "용역및공사명", "공사및용역명", "입찰건명"]
PROJECT_NAME_LABELS = ["사업명", "사업지명"]
_NEXT_LABEL_CUT = r"(?=\s*(?:[가-힣]|\d+)\s*[.)]\s*[가-힣\s]{1,8}\s*[:：])"


def _fuzzy(label: str) -> str:
    return r"\s*".join(re.escape(ch) for ch in label)


def _grab_label_value(flat: str, labels: list, max_len: int = 120) -> str:
    for lab in labels:
        m = re.search(_fuzzy(lab) + r"\s*[:：]\s*", flat)
        if m:
            seg = flat[m.end():m.end() + max_len]
            cut = re.search(_NEXT_LABEL_CUT, seg)
            if cut:
                seg = seg[:cut.start()]
            return re.sub(r"\s+", " ", seg).strip(" .,·-")
    return ""


def _title_after_notice_no(flat: str) -> str:
    m = re.search(r"공\s*고\s*제?\s*[0-9A-Za-z\s\-–—]{1,20}?호", flat)
    if not m:
        return ""
    seg = flat[m.end():m.end() + 90]
    m2 = re.search(r"입찰\s*(재)?공고", seg)
    if m2:
        seg = seg[:m2.end()]
    return re.sub(r"\s+", " ", seg).strip(" .,·-")


def extract_service_name(text: str):
    if not text:
        return "", ""
    flat = re.sub(r"\s+", " ", text)
    svc = _grab_label_value(flat, SERVICE_NAME_LABELS)
    if svc:
        proj = _grab_label_value(flat, PROJECT_NAME_LABELS)
        if proj and proj in svc:
            svc = svc.replace(proj, "").strip(" .,·-")
        if 4 <= len(svc) <= 80:
            return svc, "용역명"
    title = _title_after_notice_no(flat)
    if title and 4 <= len(title) <= 80:
        return title, "공고제목"
    return "", ""


def classify_from_document(text: str, exclude_title_keywords: list):
    """공고문 본문 기반 분류. (구분, 세부내역, 용역명) 반환. 실패하면 ("기타", "", "")."""
    if not text:
        return "기타", "", ""
    flat = re.sub(r"\s+", " ", text)

    for getter in (
        lambda: _grab_label_value(flat, SERVICE_NAME_LABELS),
        lambda: _title_after_notice_no(flat),
    ):
        name = getter()
        if not name or not (4 <= len(name) <= 80):
            continue
        proj = _grab_label_value(flat, PROJECT_NAME_LABELS)
        if proj and proj in name:
            name = name.replace(proj, "").strip(" .,·-")
        if title_excluded(name, exclude_title_keywords):
            continue
        cat, det = classify_title(name)
        if cat != "기타":
            return cat, det, name
    return "기타", "", ""


def extract_location_snippet(text: str) -> str:
    m = re.search(r"(사업지|사업)?\s*위\s*치\s*[:：]?", text)
    if m:
        start = m.end()
        return text[start:start + 80]
    return ""


REGION_NAME_NORMALIZE = {
    "서울특별시": "서울시",
    "부산광역시": "부산시",
    "대구광역시": "대구시",
    "인천광역시": "인천시",
    "광주광역시": "광주시",
    "대전광역시": "대전시",
    "울산광역시": "울산시",
    "세종특별자치시": "세종시",
}

SIDO_NAMES = [
    "서울특별시", "부산광역시", "대구광역시", "인천광역시", "광주광역시", "대전광역시",
    "울산광역시", "세종특별자치시", "경기도", "강원특별자치도", "강원도",
    "충청북도", "충청남도", "전북특별자치도", "전라북도", "전라남도",
    "경상북도", "경상남도", "제주특별자치도",
    "서울시", "부산시", "대구시", "인천시", "광주시", "대전시", "울산시", "세종시",
]

NON_REGION_WORDS = {
    "도시", "실시", "고시", "게시", "표시", "제시", "전시", "임시", "동시", "즉시",
    "수시", "다시", "일시", "당시", "명시", "공시", "적시", "명세시", "무시", "중시",
}


def normalize_region_name(region: str) -> str:
    for full, short in REGION_NAME_NORMALIZE.items():
        if region.startswith(full):
            return region.replace(full, short, 1)
    return region


def find_region_in_text(text: str) -> str:
    snippet = extract_location_snippet(text)
    search_targets = [snippet, text] if snippet else [text]

    for target in search_targets:
        m = re.search(
            r"(" + "|".join(SIDO_NAMES) + r")"
            r"(\s*[가-힣]{1,6}(?:시|군|구)(?![가-힣]))?"
            r"(\s*[가-힣0-9]{1,8}(?:구|읍|면)(?![가-힣]))?",
            target,
        )
        if m:
            parts = [p.strip() for p in m.groups() if p]
            if parts:
                return normalize_region_name(" ".join(parts))

        m = re.search(r"([가-힣]{2,6}(?:시|군)(?![가-힣]))\s*([가-힣0-9]{1,8}(?:구|읍|면)(?![가-힣]))", target)
        if m and m.group(1) not in NON_REGION_WORDS:
            return normalize_region_name(f"{m.group(1)} {m.group(2)}")
    return ""


AREA_LABELS = {
    "구역면적(㎡)": ["구역면적", "정비구역면적", "사업구역면적"],
    "대지면적(㎡)": ["대지면적", "부지면적", "사업부지면적"],
    "건축연면적(㎡)": ["건축연면적", "신축연면적", "연면적", "총연면적"],
}

SQM_TO_PYEONG = 3.305785


def _fuzzy_label_pattern(label: str) -> str:
    return r"\s*".join(re.escape(ch) for ch in label)


def format_area_number(value: float) -> str:
    return f"{round(value):,}"


def extract_area_info(text: str) -> dict:
    result = {}
    for col_m2, labels in AREA_LABELS.items():
        for label in labels:
            pattern = _fuzzy_label_pattern(label) + r"\s*[:：\(]?\s*([\d,]+\.?\d*)\s*(?:제곱미터|㎡|m2|m²)"
            m = re.search(pattern, text)
            if m:
                value_m2 = float(m.group(1).replace(",", ""))
                col_py = col_m2.replace("(㎡)", "(평)")
                result[col_m2] = format_area_number(value_m2)
                result[col_py] = format_area_number(value_m2 / SQM_TO_PYEONG)
                break
    return result


SITE_BRIEFING_LABELS = [
    "현장설명회", "현장 설명회", "현장설명일", "현장설명일자", "현장 설명일",
    "현장설명회 개최", "현장답사", "현장 답사",
]


def extract_site_briefing_date(text: str) -> str:
    """공고문/지침서 본문에서 '현장설명회' 날짜를 찾아 YYYY-MM-DD로 반환. 못 찾으면 빈 문자열."""
    if not text:
        return ""
    for label in SITE_BRIEFING_LABELS:
        m = re.search(
            _fuzzy_label_pattern(label) + r"[^0-9]{0,15}(\d{4})[.\-년]\s?(\d{1,2})[.\-월]\s?(\d{1,2})",
            text,
        )
        if m:
            year, month, day = m.group(1), m.group(2), m.group(3)
            try:
                return f"{year}-{int(month):02d}-{int(day):02d}"
            except ValueError:
                continue
    return ""


def download_attachments_for_row(row, attachment_dir: Path):
    """
    한 공고(row)에 대해 원본 응답 안의 URL들을 다운로드 시도.
    반환값: (처리 결과 메모, 문서에서 찾은 지역, 면적 정보 dict, 합친 문서 본문 텍스트)
    """
    item = row.get("_raw_dict", {})
    if not isinstance(item, dict):
        return "원본 데이터 없음", "", {}, ""

    urls = find_urls_in_item(item)
    if not urls:
        return "다운로드 링크 없음(URL 필드 미발견)", "", {}, ""

    institution = row.get("발주기관", "")
    project_type = row.get("_사업유형", "")
    content_label = row.get("_세부내역") or row.get("_구분") or ""

    folder_name = sanitize_filename(f"{institution}_{row['공고명']}")
    folder_path = Path(attachment_dir) / folder_name
    folder_path.mkdir(parents=True, exist_ok=True)

    doc_labels = ["공고문", "지침서"]
    results = []
    region_found = ""
    area_found = {}
    doc_texts = []

    for idx, (field_name, url) in enumerate(urls, start=1):
        download_url = url
        tmp_path = folder_path / f"_tmp_dl_{idx}.dat"

        ok = _download_file(download_url, str(tmp_path), timeout=30)
        if not ok and download_url.startswith("http://"):
            download_url = "https://" + download_url[len("http://"):]
            ok = _download_file(download_url, str(tmp_path), timeout=30)

        if not ok:
            results.append(f"{field_name} 실패(다운로드 오류)")
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except OSError:
                    pass
            continue

        content_head = tmp_path.read_bytes()[:8]
        content_head_long = tmp_path.read_bytes()[:300].lstrip().lower()
        looks_like_html_error = (
            content_head_long.startswith(b"<!doctype html")
            or content_head_long.startswith(b"<html")
            or content_head_long.startswith(b"<?xml")  # 일부 오류/안내 응답이 xml로 오는 경우
            or b"<html" in content_head_long[:100]
        )
        # 실제 문서(매직바이트로 인식됨)인데 우연히 <html 문자열을 포함하는 경우는 오탐 방지
        is_recognized_doc = (
            content_head.startswith(b"%PDF")
            or content_head.startswith(b"PK\x03\x04")
            or content_head.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1")
        )
        if looks_like_html_error and not is_recognized_doc:
            results.append(f"{field_name} 실패(서버가 오류/안내 페이지를 반환함 - 문서 아님)")
            try:
                tmp_path.unlink()
            except OSError:
                pass
            continue

        ext = guess_extension(download_url, "", content_head)
        doc_label = doc_labels[idx - 1] if idx <= len(doc_labels) else f"첨부{idx}"

        name_parts = [p for p in [institution, project_type, content_label, doc_label] if p]
        file_name = sanitize_filename(" ".join(name_parts)) + ext
        file_path = folder_path / file_name

        try:
            tmp_path.replace(file_path)
            results.append(f"{field_name} -> {file_name}")
        except OSError as e:
            results.append(f"{field_name} 저장실패({e})")
            continue

        doc_text = extract_text_from_file(str(file_path))
        if doc_text:
            doc_texts.append(doc_text)
        if not region_found:
            region_found = find_region_in_text(doc_text)
        if not area_found:
            area_found = extract_area_info(doc_text)

    return "; ".join(results), region_found, area_found, "\n".join(doc_texts)


# ---------------------------- 엑셀 출력 ----------------------------

FULL_COLS = [
    "폴더명", "연번", "구분", "세부내역", "공고일시", "마감일시", "현장설명회일", "지역", "공고명", "발주기관", "업체명",
    "건축연면적(㎡)", "건축연면적(평)", "대지면적(㎡)", "대지면적(평)",
    "구역면적(㎡)", "구역면적(평)", "입찰금액(원)", "평단가(원)", "비고",
]

HEADER_LABELS = {
    "건축연면적(㎡)": "건축연면적\n(㎡)",
    "건축연면적(평)": "건축연면적\n(평)",
    "대지면적(㎡)": "대지면적\n(㎡)",
    "대지면적(평)": "대지면적\n(평)",
    "구역면적(㎡)": "구역면적\n(㎡)",
    "구역면적(평)": "구역면적\n(평)",
    "입찰금액(원)": "입찰금액\n(원)",
    "평단가(원)": "평단가\n(원)",
}

COL_WIDTHS = {
    "폴더명": 11, "연번": 6, "구분": 10, "세부내역": 14, "공고일시": 17, "마감일시": 17, "지역": 14,
    "공고명": 42, "발주기관": 30, "업체명": 22,
    "건축연면적(㎡)": 10, "건축연면적(평)": 10, "대지면적(㎡)": 10, "대지면적(평)": 10,
    "구역면적(㎡)": 10, "구역면적(평)": 10, "입찰금액(원)": 13, "평단가(원)": 12, "비고": 20,
}

HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="맑은 고딕", bold=True, size=13, color="1F3864")
BODY_FONT = Font(name="맑은 고딕", size=10)
UNCLEAR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="9C5700")
UNCLEAR_FILL = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_excel(ws, data_row_count: int):
    n_cols = len(FULL_COLS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.row_dimensions[2].height = 32
    for col_idx, col_name in enumerate(FULL_COLS, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = HEADER_LABELS.get(col_name, col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for r in range(3, 3 + data_row_count):
        ws.row_dimensions[r].height = 34
        note_value = ws.cell(row=r, column=FULL_COLS.index("비고") + 1).value
        is_unclear = note_value and "수동확인" in str(note_value)
        for col_idx, col_name in enumerate(FULL_COLS, start=1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER
            if is_unclear:
                cell.fill = UNCLEAR_FILL
                cell.font = UNCLEAR_FONT
            else:
                cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(FULL_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 15)

    ws.freeze_panes = "A3"


def _write_excel(df_out: pd.DataFrame, excel_path: Path, title_str: str):
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(str(excel_path), engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, header=True, startrow=1, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]
        ws.cell(row=1, column=1, value=f"■ 정비사업 입찰공고문 ({title_str})")
        style_excel(ws, len(df_out))


# ---------------------------- 결과 객체 ----------------------------

@dataclass
class PipelineResult:
    ok: bool
    message: str = ""
    title_date: Optional[datetime] = None
    raw_count: int = 0
    filtered_df: pd.DataFrame = field(default_factory=pd.DataFrame)
    excel_path: Optional[str] = None
    attachment_dir: Optional[str] = None


# ---------------------------- 오늘의 공고 수집 ----------------------------

def run_pipeline(
    service_key: str,
    cfg: dict,
    yesterday_only: bool = True,
    days_back: int = 1,
    download_attachments: bool = True,
    progress_cb: ProgressCB = None,
    start_date=None,
    end_date=None,
) -> PipelineResult:
    """
    조회 기간은 다음 세 가지 모드 중 하나로 결정됩니다 (우선순위 순):
      1) start_date와 end_date를 둘 다 넘기면: 그 날짜 범위(포함) 전체를 조회합니다.
      2) yesterday_only=True (기본값): 직전 영업일만 조회합니다.
      3) yesterday_only=False: days_back일 전부터 지금까지 조회합니다.
    """
    logger = _Logger("today", progress_cb)
    try:
        if not service_key:
            logger.close()
            return PipelineResult(ok=False, message="SERVICE_KEY가 설정되지 않았습니다.")

        include_institution_keywords = cfg.get("include_institution_keywords", DEFAULT_INCLUDE_INSTITUTION_KEYWORDS)
        include_title_keywords = cfg.get("include_title_keywords", DEFAULT_INCLUDE_TITLE_KEYWORDS)
        exclude_title_keywords = cfg.get("exclude_title_keywords", DEFAULT_EXCLUDE_TITLE_KEYWORDS)

        if start_date is not None and end_date is not None:
            start_dt = datetime.combine(start_date, datetime.min.time())
            end_dt = datetime.combine(end_date, datetime.min.time()) + timedelta(days=1, minutes=-1)
            if end_dt < start_dt:
                logger.close()
                return PipelineResult(ok=False, message="종료일이 시작일보다 앞설 수 없습니다.")
            date_chunks = _make_date_chunks_for_range(start_dt, end_dt)
            target_day = start_dt
            restrict_to_date = None
        elif yesterday_only:
            target_day, date_chunks = _make_date_chunks_for_yesterday(look_ahead_for_cancel=True)
            restrict_to_date = target_day.strftime("%Y-%m-%d")
        else:
            end_dt = datetime.now()
            start_dt = end_dt - timedelta(days=days_back)
            date_chunks = _make_date_chunks_for_range(start_dt, end_dt)
            target_day = start_dt
            restrict_to_date = None

        logger.log("공고 조회를 시작합니다...")
        raw_items = collect_raw(["용역"], date_chunks, service_key, logger=logger)
        if not raw_items:
            logger.close()
            return PipelineResult(
                ok=False, message="API 조회 결과가 없습니다. (기간을 확인해주세요)",
                title_date=target_day, raw_count=0,
            )

        logger.log(f"조회 완료: 총 {len(raw_items)}건. 필터링을 시작합니다...")
        df = apply_filters_and_classify(
            raw_items,
            include_institution_keywords,
            include_title_keywords,
            exclude_title_keywords,
            restrict_to_date=restrict_to_date,
            logger=logger,
        )
        if df.empty:
            logger.close()
            return PipelineResult(
                ok=False,
                message=f"조회된 {len(raw_items)}건 중 필터를 통과한 공고가 없습니다.",
                title_date=target_day, raw_count=len(raw_items),
            )

        df = df.reset_index(drop=True)
        df.insert(0, "연번", range(1, len(df) + 1))

        attachment_dir = DEFAULT_ATTACHMENT_DIR
        if download_attachments:
            logger.log(f"첨부파일(공고문/지침서) 다운로드를 시작합니다 ({len(df)}건)...")
            if attachment_dir.exists():
                shutil.rmtree(attachment_dir)  # 이전 실행분 폴더가 계속 쌓이는 것 방지
            attachment_dir.mkdir(parents=True, exist_ok=True)
            reclassified = 0
            for idx, row in df.iterrows():
                result_note, region_from_doc, area_info, doc_text = download_attachments_for_row(row, attachment_dir)
                if region_from_doc:
                    df.at[idx, "지역"] = region_from_doc
                for col_name, value in area_info.items():
                    df.at[idx, col_name] = value
                if doc_text:
                    briefing_date = extract_site_briefing_date(doc_text)
                    if briefing_date:
                        df.at[idx, "현장설명회일"] = briefing_date

                if df.at[idx, "구분"] == "기타" and "협력업체" in str(row["공고명"]) and doc_text:
                    cat, det, svc_name = classify_from_document(doc_text, exclude_title_keywords)
                    if cat != "기타":
                        df.at[idx, "구분"] = cat
                        df.at[idx, "세부내역"] = det
                        if svc_name and svc_name not in str(row["공고명"]):
                            df.at[idx, "공고명"] = f"{row['공고명']} ({svc_name})"
                        old = str(df.at[idx, "비고"] or "")
                        df.at[idx, "비고"] = (old + f" [문서추출/{cat}/{det}]").strip()
                        reclassified += 1
                logger.log(f"  - {row['공고명']}: {result_note}")
            if reclassified:
                logger.log(f"문서 기반 재분류: {reclassified}건")

        df_out = df[FULL_COLS].copy()
        df_out["연번"] = range(1, len(df_out) + 1)

        if start_date is not None and end_date is not None:
            title_str = f"{start_date.strftime('%Y-%m-%d')}~{end_date.strftime('%Y-%m-%d')}"
        else:
            title_str = target_day.strftime("%Y-%m-%d")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = OUTPUT_DIR / f"정비사업입찰공고_{title_str}_{ts}.xlsx"
        _write_excel(df_out, excel_path, title_str)

        logger.log(f"완료: 필터 통과 {len(df_out)}건 -> 엑셀 저장 완료")
        logger.close()
        return PipelineResult(
            ok=True,
            message="완료",
            title_date=target_day,
            raw_count=len(raw_items),
            filtered_df=df_out,
            excel_path=str(excel_path),
            attachment_dir=str(attachment_dir) if download_attachments else None,
        )
    except Exception as e:
        logger.log(f"[!] 오류: {type(e).__name__}: {e}")
        logger.close()
        return PipelineResult(ok=False, message=f"오류가 발생했습니다: {e}")


# ---------------------------- 조합별 이력 조회 ----------------------------

def search_by_institution(
    institution_keyword: str,
    service_key: str,
    cfg: dict,
    years_back: float = 1.0,
    biz_types: Optional[list] = None,
    download_attachments: bool = True,
    progress_cb: ProgressCB = None,
) -> PipelineResult:
    logger = _Logger("hist", progress_cb)
    try:
        if not service_key:
            logger.close()
            return PipelineResult(ok=False, message="SERVICE_KEY가 설정되지 않았습니다.")
        if not institution_keyword or not institution_keyword.strip():
            logger.close()
            return PipelineResult(ok=False, message="발주기관 이름을 입력해주세요.")

        biz_types = biz_types or ["용역"]
        include_title_keywords = cfg.get("include_title_keywords", DEFAULT_INCLUDE_TITLE_KEYWORDS)
        exclude_title_keywords = cfg.get("exclude_title_keywords", DEFAULT_EXCLUDE_TITLE_KEYWORDS)

        keyword = institution_keyword.strip()

        def institution_match_fn(inst: str) -> bool:
            return keyword in inst

        end_dt = datetime.now()
        start_dt = end_dt - timedelta(days=int(round(years_back * 365)))
        date_chunks = _make_date_chunks_for_range(start_dt, end_dt)

        est_calls = len(date_chunks) * len(biz_types)
        logger.log(f"'{keyword}' 관련 공고를 최근 {years_back}년치 조회합니다 (예상 {est_calls}회 호출)...")

        raw_items = collect_raw(biz_types, date_chunks, service_key, logger=logger)
        if not raw_items:
            logger.close()
            return PipelineResult(ok=False, message="API 조회 결과가 없습니다.", raw_count=0)

        logger.log(f"조회 완료: 총 {len(raw_items)}건. 필터링을 시작합니다...")
        df = apply_filters_and_classify(
            raw_items,
            include_institution_keywords=[],
            include_title_keywords=include_title_keywords,
            exclude_title_keywords=exclude_title_keywords,
            institution_match_fn=institution_match_fn,
            restrict_to_date=None,
            logger=logger,
        )
        if df.empty:
            logger.close()
            return PipelineResult(
                ok=False,
                message=f"'{keyword}' 관련 공고 {len(raw_items)}건 중 필터를 통과한 공고가 없습니다.",
                raw_count=len(raw_items),
            )

        df = df.sort_values("공고일시", ascending=False).reset_index(drop=True)
        df.insert(0, "연번", range(1, len(df) + 1))

        attachment_dir = OUTPUT_DIR / f"이력조회_{sanitize_filename(keyword)}_첨부파일"
        if download_attachments:
            logger.log(f"첨부파일 다운로드를 시작합니다 ({len(df)}건)...")
            attachment_dir.mkdir(parents=True, exist_ok=True)
            for idx, row in df.iterrows():
                result_note, region_from_doc, area_info, doc_text = download_attachments_for_row(row, attachment_dir)
                if region_from_doc:
                    df.at[idx, "지역"] = region_from_doc
                for col_name, value in area_info.items():
                    df.at[idx, col_name] = value
                if doc_text:
                    briefing_date = extract_site_briefing_date(doc_text)
                    if briefing_date:
                        df.at[idx, "현장설명회일"] = briefing_date
                logger.log(f"  - {row['공고명']}: {result_note}")

        df_out = df[FULL_COLS].copy()
        df_out["연번"] = range(1, len(df_out) + 1)

        title_str = f"{keyword} 이력조회"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = OUTPUT_DIR / f"이력조회_{sanitize_filename(keyword)}_{ts}.xlsx"
        _write_excel(df_out, excel_path, title_str)

        logger.log(f"완료: 필터 통과 {len(df_out)}건 -> 엑셀 저장 완료")
        logger.close()
        return PipelineResult(
            ok=True,
            message="완료",
            raw_count=len(raw_items),
            filtered_df=df_out,
            excel_path=str(excel_path),
            attachment_dir=str(attachment_dir) if download_attachments else None,
        )
    except Exception as e:
        logger.log(f"[!] 오류: {type(e).__name__}: {e}")
        logger.close()
        return PipelineResult(ok=False, message=f"오류가 발생했습니다: {e}")
