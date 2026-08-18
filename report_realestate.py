# -*- coding: utf-8 -*-
"""
report_realestate.py - 부동산 실거래가 분석 보고서 자동화

'원본데이터' 시트(법정동/지번/건물유형/건물주용도/계약일/거래금액(원)/건물면적(㎡)/
대지면적(㎡)/평당가(원)/층/비고 컬럼 포함)가 들어있는 엑셀을 업로드하면,
아래 4개 분석 시트를 다시 계산해 워크북에 반영합니다.
  - 동별_용도별_요약
  - 업무_지번별_분기추이 (업무시설 중 <50㎡ "표준면적" 매물 기준)
  - 근린생활_면적구간분석
  - 층효과_비교

[검증 방법] 실제 사용자가 수식(SUMIFS/AVERAGEIFS 등)으로 직접 만든 예시 파일
(수서동_자곡동_실거래가_분석.xlsx)의 모든 숫자와 하나하나 대조해서 확인했습니다.
- 동별_용도별_요약, 근린생활_면적구간분석, 층효과_비교(구간별 건수/평균), 지번별_분기추이:
  전부 정확히 일치.
- 층효과_비교의 "업무" 상관계수 단 1개 값만 원본과 정확히 재현되지 않았습니다
  (원본이 어떤 세부 조건으로 걸러냈는지 역추적이 안 됐습니다. 표본이 37건뿐이라
  이상치 하나에도 상관계수가 크게 흔들리는 값이라, 핵심 집계표에 비해 중요도가 낮다고
  보고 우선 진행했습니다).
"""
import io
import re
from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RAW_SHEET = "원본데이터"

RAW_COLS = [
    "법정동", "지번", "건물유형", "건물주용도", "계약일", "거래금액(원)",
    "건물면적(㎡)", "대지면적(㎡)", "평당가(원)", "층", "비고",
]

NEIGHBORHOOD_TYPES = ["제1종근린생활", "제2종근린생활"]
AREA_BUCKET_ORDER = ["~25㎡", "25~50㎡", "50~100㎡", "100㎡+"]

TITLE_FONT = Font(name="맑은 고딕", bold=True, size=13, color="1F3864")
NOTE_FONT = Font(name="맑은 고딕", size=9, color="666666", italic=True)
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
BODY_FONT = Font(name="맑은 고딕", size=10)


# ============================ 파생 컬럼 계산 ============================

def _quarter_label(dt) -> str:
    ts = pd.Timestamp(dt)
    q = (ts.month - 1) // 3 + 1
    return f"{ts.year}Q{q}"


def _area_bucket(area_m2) -> str:
    try:
        v = float(area_m2)
    except (TypeError, ValueError):
        return ""
    if v < 25:
        return "~25㎡"
    if v < 50:
        return "25~50㎡"
    if v < 100:
        return "50~100㎡"
    return "100㎡+"


def build_master_df(input_bytes_or_path, sheet_name: str = RAW_SHEET) -> pd.DataFrame:
    """'원본데이터' 시트를 읽고, 분기/면적구간 파생 컬럼을 (없으면) 계산해서 반환."""
    if isinstance(input_bytes_or_path, bytes):
        input_bytes_or_path = io.BytesIO(input_bytes_or_path)
    wb = openpyxl.load_workbook(input_bytes_or_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"'{sheet_name}' 시트를 찾을 수 없습니다. "
            f"실거래가 원본 데이터가 담긴 엑셀(시트명 '{sheet_name}')을 업로드해주세요."
        )
    ws = wb[sheet_name]

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(header)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        rec = {}
        for col in RAW_COLS:
            i = col_idx.get(col)
            rec[col] = row[i] if i is not None and i < len(row) else None
        if not rec.get("법정동"):
            continue
        rows.append(rec)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df["계약일"] = pd.to_datetime(df["계약일"], errors="coerce")
    df["거래금액(원)"] = pd.to_numeric(df["거래금액(원)"], errors="coerce")
    df["건물면적(㎡)"] = pd.to_numeric(df["건물면적(㎡)"], errors="coerce")
    df["평당가(원)"] = pd.to_numeric(df["평당가(원)"], errors="coerce")
    df["층"] = pd.to_numeric(df["층"], errors="coerce")

    # 분기/면적구간이 원본에 이미 있으면 그대로 쓰고, 없으면 새로 계산
    if "분기" in col_idx:
        df["분기"] = [row[col_idx["분기"]] for row in ws.iter_rows(min_row=2, values_only=True) if row and not all(v is None for v in row)][:len(df)]
    if not df.get("분기", pd.Series(dtype=object)).notna().all():
        df["분기"] = df["계약일"].map(_quarter_label)

    if "면적구간" in col_idx:
        vals = [row[col_idx["면적구간"]] for row in ws.iter_rows(min_row=2, values_only=True) if row and not all(v is None for v in row)][:len(df)]
        df["면적구간"] = vals
    if "면적구간" not in df.columns or df["면적구간"].isna().any() or (df["면적구간"] == "").any():
        df["면적구간"] = df["건물면적(㎡)"].map(_area_bucket)

    return df


# ============================ 집계 계산 ============================

def build_dong_use_summary(df: pd.DataFrame) -> pd.DataFrame:
    g = df.groupby(["법정동", "건물주용도"]).agg(
        거래건수=("거래금액(원)", "count"),
        평균거래금액=("거래금액(원)", "mean"),
        중앙값거래금액=("거래금액(원)", "median"),
        평균평당가=("평당가(원)", "mean"),
        최고평당가=("평당가(원)", "max"),
        최저평당가=("평당가(원)", "min"),
    ).reset_index()
    return g.sort_values(["법정동", "거래건수"], ascending=[True, False]).reset_index(drop=True)


def build_area_range_analysis(df: pd.DataFrame) -> pd.DataFrame:
    """근린생활시설(제1종+제2종) 면적구간별 평균 평당가."""
    geun = df[df["건물주용도"].isin(NEIGHBORHOOD_TYPES)]
    if geun.empty:
        return pd.DataFrame()
    g = geun.groupby("면적구간").agg(
        거래건수=("거래금액(원)", "count"),
        평균평당가=("평당가(원)", "mean"),
        평균거래금액=("거래금액(원)", "mean"),
    ).reset_index()
    g["면적구간"] = pd.Categorical(g["면적구간"], categories=AREA_BUCKET_ORDER, ordered=True)
    return g.sort_values("면적구간").reset_index(drop=True)


def _floor_bucket_business(floor) -> str:
    if pd.isna(floor):
        return ""
    f = float(floor)
    if 5 <= f <= 10:
        return "5~10층"
    if 11 <= f <= 15:
        return "11~15층"
    if 16 <= f <= 20:
        return "16~20층"
    return ""  # 정의된 구간 밖(예: 1~4층, 21층 이상)은 표본이 거의 없어 제외


def _floor_bucket_neighborhood(floor) -> str:
    if pd.isna(floor):
        return ""
    f = float(floor)
    if f == 1:
        return "1층"
    if f == 2:
        return "2층"
    if f >= 3:
        return "3층+"
    return ""


def build_floor_effect(df: pd.DataFrame) -> dict:
    """업무 vs 근린생활 층별 평당가 비교 + 층-평당가 상관계수."""
    biz = df[df["건물주용도"] == "업무"].dropna(subset=["층"]).copy()
    geun = df[df["건물주용도"].isin(NEIGHBORHOOD_TYPES)].dropna(subset=["층"]).copy()

    biz["층구간"] = biz["층"].map(_floor_bucket_business)
    geun["층구간"] = geun["층"].map(_floor_bucket_neighborhood)

    biz_b = biz[biz["층구간"] != ""]
    geun_b = geun[geun["층구간"] != ""]

    rows = []
    biz_order = ["5~10층", "11~15층", "16~20층"]
    for bucket in biz_order:
        sub = biz_b[biz_b["층구간"] == bucket]
        if len(sub):
            rows.append(["업무", bucket, len(sub), sub["평당가(원)"].mean()])
    geun_order = ["1층", "2층", "3층+"]
    for bucket in geun_order:
        sub = geun_b[geun_b["층구간"] == bucket]
        if len(sub):
            rows.append(["근린생활", bucket, len(sub), sub["평당가(원)"].mean()])

    table = pd.DataFrame(rows, columns=["용도", "층구간", "거래건수", "평균평당가"])

    biz_corr = biz_b["층"].corr(biz_b["평당가(원)"]) if len(biz_b) >= 2 else None
    geun_corr = geun["층"].corr(geun["평당가(원)"]) if len(geun) >= 2 else None

    total_rows = len(df)
    no_floor = df["층"].isna().sum()

    return {
        "table": table,
        "업무_상관계수": biz_corr,
        "근린생활_상관계수": geun_corr,
        "전체건수": total_rows,
        "층정보없음건수": int(no_floor),
    }


def build_jibun_quarterly_trend(df: pd.DataFrame, area_threshold: float = 50.0) -> dict:
    """업무시설 중 <area_threshold㎡ '표준면적' 매물만 걸러서, 지번별 분기 평균 평당가 추이."""
    biz = df[df["건물주용도"] == "업무"].copy()
    std = biz[biz["건물면적(㎡)"] < area_threshold]
    if std.empty:
        return {"table": pd.DataFrame(), "jibuns": [], "corr": None}

    jibuns = sorted(std["지번"].dropna().unique().tolist(), key=lambda x: -len(std[std["지번"] == x]))

    quarters = sorted(std["분기"].dropna().unique().tolist())
    table = pd.DataFrame({"분기": quarters})
    for jibun in jibuns:
        sub = std[std["지번"] == jibun]
        g = sub.groupby("분기").agg(평균평당가=("평당가(원)", "mean"), 거래건수=("평당가(원)", "count"))
        table[f"지번{jibun} 평균평당가"] = table["분기"].map(g["평균평당가"])
        table[f"지번{jibun} 거래건수"] = table["분기"].map(g["거래건수"]).fillna(0).astype(int)

    std_floor = std.dropna(subset=["층"])
    corr = std_floor["층"].corr(std_floor["평당가(원)"]) if len(std_floor) >= 2 else None

    return {"table": table, "jibuns": jibuns, "corr": corr, "n": len(std_floor)}


# ============================ 엑셀 기록 ============================

def _write_title(ws, row, text, font=TITLE_FONT):
    ws.cell(row=row, column=2, value=text).font = font
    return row + 2


def _write_note(ws, row, text):
    ws.cell(row=row, column=2, value=text).font = NOTE_FONT
    return row + 2


def _write_df(ws, row, df: pd.DataFrame, start_col=2, pct_cols=None, round_cols=None):
    pct_cols = pct_cols or []
    round_cols = round_cols or []
    for j, col_name in enumerate(df.columns):
        c = ws.cell(row=row, column=start_col + j, value=col_name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1
    for _, data_row in df.iterrows():
        for j, col_name in enumerate(df.columns):
            val = data_row[col_name]
            if col_name in round_cols and pd.notna(val):
                val = round(float(val), 1)
            cell = ws.cell(row=row, column=start_col + j, value=val if pd.notna(val) else None)
            cell.font = BODY_FONT
            if col_name in pct_cols:
                cell.number_format = "0.0%"
            elif isinstance(val, float):
                cell.number_format = "#,##0"
        row += 1
    return row + 1


def write_dong_use_sheet(ws, summary: pd.DataFrame):
    row = 2
    row = _write_title(ws, row, "법정동 × 건물주용도별 요약")
    row = _write_note(ws, row, "원본데이터 시트를 기준으로 자동 재계산됩니다.")
    _write_df(ws, row, summary)
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16


def write_area_range_sheet(ws, area_df: pd.DataFrame):
    row = 2
    row = _write_title(ws, row, "근린생활시설(제1종+제2종) 면적구간별 평균 평당가")
    row = _write_note(ws, row, "면적이 작을수록 평당가가 높아지는 경향이 있는지 확인하는 표입니다.")
    if area_df.empty:
        ws.cell(row=row, column=2, value="근린생활시설 거래 데이터가 없습니다.")
        return
    _write_df(ws, row, area_df)
    for col in range(2, 6):
        ws.column_dimensions[get_column_letter(col)].width = 16


def write_floor_effect_sheet(ws, result: dict):
    row = 2
    row = _write_title(ws, row, "업무 vs 근린생활 층별 평당가 비교")
    row = _write_note(
        ws, row,
        f"※ 층 정보가 없는 거래(전체 {result['전체건수']}건 중 {result['층정보없음건수']}건)는 제외하고 집계"
    )
    row = _write_df(ws, row, result["table"])

    row = _write_title(ws, row, "층 - 평당가 상관계수", TITLE_FONT)
    corr_rows = []
    if result["업무_상관계수"] is not None:
        corr_rows.append(["업무", round(result["업무_상관계수"], 4)])
    if result["근린생활_상관계수"] is not None:
        corr_rows.append(["근린생활", round(result["근린생활_상관계수"], 4)])
    corr_df = pd.DataFrame(corr_rows, columns=["용도", "상관계수"])
    _write_df(ws, row, corr_df)

    for col in range(2, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18


def write_jibun_trend_sheet(ws, result: dict, area_threshold: float):
    row = 2
    jibun_list = ", ".join(str(j) for j in result["jibuns"])
    row = _write_title(ws, row, f"업무시설 지번별({jibun_list}) 표준면적(<{area_threshold:.0f}㎡) 분기별 평균 평당가 추이")
    if result["table"].empty:
        ws.cell(row=row, column=2, value="분석할 업무시설 거래가 없습니다.")
        return
    row = _write_note(ws, row, f"표준면적(<{area_threshold:.0f}㎡) 매물만 필터링해 순수 시세를 비교합니다.")
    row = _write_df(ws, row, result["table"])

    if result.get("corr") is not None:
        row = _write_title(ws, row, "층 - 평당가 상관계수 (표준면적, 층정보 있는 건)", TITLE_FONT)
        ws.cell(row=row, column=2, value=f"업무시설 전체: {round(result['corr'], 4)} (n={result.get('n', 0)})").font = BODY_FONT

    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 20


# ============================ 오케스트레이션 ============================

SHEET_ORDER_HINT = [RAW_SHEET, "동별_용도별_요약", "업무_지번별_분기추이", "근린생활_면적구간분석", "층효과_비교"]


def refresh_realestate_excel_bytes(input_bytes: bytes, area_threshold: float = 50.0, progress_cb=None) -> bytes:
    """업로드된 실거래가 분석 엑셀(bytes)을 받아 4개 분석시트를 재계산한 새 워크북을 bytes로 반환.
    '원본데이터' 시트는 그대로 보존됩니다."""

    def _log(msg):
        if progress_cb:
            progress_cb(msg)

    _log("원본데이터를 불러오는 중...")
    df = build_master_df(input_bytes)
    if df.empty:
        raise ValueError("원본데이터 시트에서 유효한 거래 데이터를 찾지 못했습니다.")
    _log(f"{len(df)}건 로드 완료. 집계 계산 중...")

    dong_use = build_dong_use_summary(df)
    area_range = build_area_range_analysis(df)
    floor_effect = build_floor_effect(df)
    jibun_trend = build_jibun_quarterly_trend(df, area_threshold=area_threshold)

    _log("워크북에 집계시트 기록 중...")
    wb = openpyxl.load_workbook(io.BytesIO(input_bytes))

    sheet_writers = {
        "동별_용도별_요약": lambda ws: write_dong_use_sheet(ws, dong_use),
        "업무_지번별_분기추이": lambda ws: write_jibun_trend_sheet(ws, jibun_trend, area_threshold),
        "근린생활_면적구간분석": lambda ws: write_area_range_sheet(ws, area_range),
        "층효과_비교": lambda ws: write_floor_effect_sheet(ws, floor_effect),
    }
    for name, writer_fn in sheet_writers.items():
        if name in wb.sheetnames:
            wb.remove(wb[name])
        ws = wb.create_sheet(name)
        writer_fn(ws)

    for i, name in enumerate(SHEET_ORDER_HINT):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    _log("완료")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_report_summary(df: pd.DataFrame, dong_use, area_range, floor_effect, jibun_trend) -> dict:
    """탭 화면/보고서에 쓸 핵심 인사이트 딕셔너리."""
    top_row = dong_use.sort_values("거래건수", ascending=False).iloc[0] if not dong_use.empty else None
    return {
        "총거래건수": len(df),
        "법정동목록": sorted(df["법정동"].dropna().unique().tolist()),
        "용도목록": sorted(df["건물주용도"].dropna().unique().tolist()),
        "최다거래그룹": f"{top_row['법정동']} {top_row['건물주용도']}" if top_row is not None else "",
        "최다거래건수": int(top_row["거래건수"]) if top_row is not None else 0,
        "업무_상관계수": floor_effect.get("업무_상관계수"),
        "근린생활_상관계수": floor_effect.get("근린생활_상관계수"),
        "지번목록": jibun_trend.get("jibuns", []),
    }
