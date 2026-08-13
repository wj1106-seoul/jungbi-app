# -*- coding: utf-8 -*-
"""
report_core.py - 정비사업 분석 보고서(엑셀 갱신 + Word용 데이터) 핵심 로직
load_master.py + analysis.py + write_excel.py + build_report_data.py를 통합.
웹앱(app.py)에서 이 모듈의 refresh_excel_bytes() / build_report_data()를 가져다 씁니다.
"""
import io
import re
import json
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime

import openpyxl
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter



# ============================================================
# --- load_master.py ---
# ============================================================

MASTER_COLS = [
    "폴더명", "연번", "구분", "세부내역", "공고일시", "지역", "공고명", "발주기관", "업체명",
    "건축연면적(㎡)", "건축연면적(평)", "대지면적(㎡)", "대지면적(평)",
    "구역면적(㎡)", "구역면적(평)", "입찰금액(원)", "평단가(원)", "비고",
]

SIDO_KEYWORDS = [
    ("서울시", "서울시"), ("부산시", "부산시"), ("대구시", "대구시"), ("인천시", "인천시"),
    ("광주시", "광주시"), ("대전시", "대전시"), ("울산시", "울산시"), ("세종시", "세종시"),
    ("경기도", "경기도"), ("강원도", "강원도"), ("충청북도", "충청북도"), ("충청남도", "충청남도"),
    ("전라북도", "전라북도"), ("전라남도", "전라남도"), ("경상북도", "경상북도"), ("경상남도", "경상남도"),
    ("제주도", "제주도"),
    ("경남", "경상남도"), ("경북", "경상북도"), ("충남", "충청남도"), ("충북", "충청북도"),
    ("전남", "전라남도"), ("전북", "전라북도"),
    ("안산시", "경기도"), ("의정부시", "경기도"), ("군포시", "경기도"), ("용인시", "경기도"),
    ("수원시", "경기도"), ("성남시", "경기도"), ("안양시", "경기도"), ("부천시", "경기도"),
    ("고양시", "경기도"), ("남양주시", "경기도"), ("하남시", "경기도"), ("화성시", "경기도"),
    ("시흥시", "경기도"), ("의왕시", "경기도"), ("구리시", "경기도"), ("김포시", "경기도"),
    ("평택시", "경기도"), ("광명시", "경기도"), ("퇴계원", "경기도"),
    ("전주시", "전라북도"), ("정읍시", "전라북도"),
    ("여수시", "전라남도"), ("순천시", "전라남도"), ("목포시", "전라남도"),
    ("구미시", "경상북도"), ("포항시", "경상북도"), ("안동시", "경상북도"), ("상주시", "경상북도"),
    ("창원시", "경상남도"), ("김해시", "경상남도"), ("양산시", "경상남도"), ("밀양시", "경상남도"),
    ("청주시", "충청북도"),
    ("서산시", "충청남도"), ("천안시", "충청남도"),
    ("동해시", "강원도"), ("춘천시", "강원도"),
    ("제주시", "제주도"),
    ("부신시", "부산시"),  # 원본 오타 케이스도 그대로 반영(이미 원본 데이터에서 수정됨)
]

# 사업키 만들 때 자르는 기준 키워드(발주기관/공고명에서 프로젝트명만 남기기 위한 컷 지점)
PROJECT_NAME_CUT_KEYWORDS = ["재건축", "재개발", "가로주택", "정비사업", "협력업체", "설계자", "조합", "추진위"]

# 원본에서 확인된 동일 사업장 표기 통일(사업키 수식에 하드코딩된 예외)
PROJECT_NAME_ALIASES = {
    "하안5단지": "하안주공5단지",
    "흑석11구역착공후지하안": "흑석11재정비촉진구역",
}


def clean_tab(s) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    return str(s).replace("\t", "").strip()


def derive_sido(region: str) -> str:
    if not region:
        return ""
    for kw, sido in SIDO_KEYWORDS:
        if kw in region:
            return sido
    return "기타"


def derive_project_type(institution: str, title: str) -> str:
    text = clean_tab(institution) + clean_tab(title)
    if not text:
        return ""
    checks = [
        ("가로주택", "가로주택정비"),
        ("소규모재건축", "소규모재건축"),
        ("소규모재개발", "소규모재개발"),
        ("자율주택", "자율주택정비"),
        ("재정비촉진", "재개발(촉진)"),
        ("도시환경", "도시환경정비"),
        ("생활권재개발", "재개발(생활권)"),
        ("재개발", "재개발"),
        ("재건축", "재건축"),
    ]
    for kw, label in checks:
        if kw in text:
            return label
    return "기타"


def derive_orderer_type(institution: str) -> str:
    inst = clean_tab(institution)
    if not inst:
        return ""
    if "신탁" in inst:
        return "신탁사"
    if "추진위" in inst:
        return "추진위원회"
    if "지역주택" in inst:
        return "지역주택조합"
    if "조합" in inst:
        return "조합"
    return "기타"


def derive_bid_status(remark: str) -> str:
    r = clean_tab(remark)
    if "개찰x" in r:
        return "결과미공개"
    if "유찰" in r:
        return "유찰"
    if "취소" in r:
        return "취소"
    return "정상"


def derive_project_name(title: str) -> str:
    """공고명에서 컷 키워드 앞부분을 프로젝트명으로 추출 (사업키의 세 번째 조각).
    엑셀 원본 수식이 최종적으로 공백을 전부 제거하므로 동일하게 처리."""
    t = clean_tab(title)
    if not t:
        return ""
    positions = [t.find(kw) for kw in PROJECT_NAME_CUT_KEYWORDS]
    positions = [p for p in positions if p >= 0]
    if not positions:
        name = t[:15]
    else:
        name = t[:min(positions)]
    name = name.strip().replace(" ", "")
    return PROJECT_NAME_ALIASES.get(name, name)


def derive_project_key(region: str, institution: str, title: str, orderer_type: str) -> str:
    """사업키(발주기관) - 신탁사는 지역|신탁사명|프로젝트명, 그 외는 지역|발주기관명."""
    inst = clean_tab(institution)
    if not inst:
        return ""
    if orderer_type == "신탁사":
        proj = derive_project_name(title)
        if not proj:
            proj = "산본1동3구역" if "산본동" in clean_tab(region) else proj
        return f"{clean_tab(region)}|{inst.replace('㈜', '(주)')}|{proj}"
    return f"{clean_tab(region)}|{inst.replace('㈜', '(주)')}"


def derive_notice_id(notice_dt, title: str, institution: str) -> str:
    if not title:
        return ""
    try:
        date_str = pd.Timestamp(notice_dt).strftime("%Y%m%d")
    except Exception:
        date_str = ""
    return f"{date_str}_{clean_tab(title)}_{clean_tab(institution)}"


def build_master_df(xlsx_path: str, sheet_name: str = "2026 정비사업 입찰공고 정리내역") -> pd.DataFrame:
    """마스터 시트를 읽어 파생 컬럼까지 전부 파이썬으로 재계산한 DataFrame을 반환."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        row = row[:18]  # A~R 열(원본 입력 컬럼)만 사용, S열 이후는 파이썬으로 재계산
        if len(row) < 8 or not row[7]:  # H열(발주기관)이 비면 스킵(스페이서 행)
            continue
        rows.append(row)

    df = pd.DataFrame(rows, columns=MASTER_COLS)

    df["발주기관"] = df["발주기관"].map(clean_tab)
    df["공고명"] = df["공고명"].map(clean_tab)
    df["지역"] = df["지역"].map(clean_tab)
    df["업체명"] = df["업체명"].map(clean_tab)
    df["비고"] = df["비고"].map(clean_tab)

    df["시도"] = df["지역"].map(derive_sido)
    df["사업유형"] = df.apply(lambda r: derive_project_type(r["발주기관"], r["공고명"]), axis=1)
    df["발주주체"] = df["발주기관"].map(derive_orderer_type)
    df["개찰상태"] = df["비고"].map(derive_bid_status)
    df["사업키(발주기관)"] = df.apply(
        lambda r: derive_project_key(r["지역"], r["발주기관"], r["공고명"], r["발주주체"]), axis=1
    )
    df["공고ID"] = df.apply(lambda r: derive_notice_id(r["공고일시"], r["공고명"], r["발주기관"]), axis=1)

    df["공고내순번"] = df.groupby("공고ID").cumcount() + 1
    df["참여업체수"] = df.groupby("공고ID")["공고ID"].transform("count")

    # 사업장대표행: 같은 사업키에서 공고내순번==1인 첫 등장 행만 1
    is_first_of_notice = df["공고내순번"] == 1
    df["사업장대표행"] = 0
    seen_keys = set()
    rep_flags = []
    for key, is_first in zip(df["사업키(발주기관)"], is_first_of_notice):
        if is_first and key not in seen_keys:
            rep_flags.append(1)
            seen_keys.add(key)
        else:
            rep_flags.append(0)
    df["사업장대표행"] = rep_flags

    def to_amount_type(v):
        if pd.isna(v):
            return ""
        try:
            v = float(v)
        except (TypeError, ValueError):
            return ""
        return "단가" if v < 1_000_000 else "총액"

    df["금액유형"] = df["입찰금액(원)"].map(to_amount_type)

    df["공고일시"] = pd.to_datetime(df["공고일시"], errors="coerce")
    df["월"] = df["공고일시"].dt.month

    return df




# ============================================================
# --- analysis.py ---
# ============================================================

PROJECT_TYPE_GROUPS = {
    "재건축": "재건축",
    "재개발": "재개발계", "재개발(촉진)": "재개발계", "재개발(생활권)": "재개발계", "도시환경정비": "재개발계",
    "가로주택정비": "가로주택정비",
    "소규모재건축": "소규모정비계", "소규모재개발": "소규모정비계", "자율주택정비": "소규모정비계",
    "기타": "기타",
}

CATEGORY_ORDER = ["설계", "엔지니어링", "감리", "공사", "토목", "해체", "구조", "친환경", "PM", "CM", "PCM", "기타"]
# 파이프라인①의 "구분별시기패턴의 순서: 설계·엔지니어링 → 감리·토목·PM" 캡션과 동일하게,
# 후반 단계는 감리/토목/PM 세 가지만, 초반(설계·엔지니어링만)은 정확히 그 두 가지로만 구성된 경우만 인정.
LATE_STAGE_CATEGORIES = {"감리", "토목", "PM"}
EARLY_ONLY_CATEGORIES = {"설계", "엔지니어링"}
# 파이프라인①에서만 쓰는 넓은 후반범주(감리/토목/PM 외 CM/PCM/해체/공사/구조도 "후속 단계"로 봄)
LATE_STAGE_BROAD = {"감리", "토목", "PM", "CM", "PCM", "해체", "공사", "구조"}


def notice_level_df(df: pd.DataFrame) -> pd.DataFrame:
    """공고 단위(같은 공고 내 여러 응찰업체는 1건) - 공고내순번==1인 행만."""
    return df[df["공고내순번"] == 1].copy()


def project_level_df(df: pd.DataFrame) -> pd.DataFrame:
    """사업장 단위(같은 사업장의 여러 공고는 1건) - 사업장대표행==1인 행만."""
    return df[df["사업장대표행"] == 1].copy()


# ---------------- 요약 ----------------

def build_summary(df: pd.DataFrame) -> dict:
    notice = notice_level_df(df)
    cat_counts = notice["구분"].value_counts().reindex(CATEGORY_ORDER, fill_value=0)
    type_counts = notice["사업유형"].value_counts()
    status_counts = notice["개찰상태"].value_counts()
    return {
        "구분별": cat_counts,
        "사업유형별": type_counts,
        "개찰상태별": status_counts,
        "총공고수": len(notice),
    }


# ---------------- 지역유형추이 ----------------

def build_region_type_trend(df: pd.DataFrame) -> pd.DataFrame:
    proj = project_level_df(df)
    proj = proj.copy()
    proj["유형그룹"] = proj["사업유형"].map(lambda x: PROJECT_TYPE_GROUPS.get(x, "기타"))
    group_cols = ["재건축", "재개발계", "가로주택정비", "소규모정비계", "기타"]
    pivot = proj.pivot_table(index="월", columns="유형그룹", values="사업키(발주기관)", aggfunc="count", fill_value=0)
    pivot = pivot.reindex(columns=group_cols, fill_value=0)
    pivot = pivot.reindex(range(1, 13), fill_value=0)
    pivot["사업장수"] = pivot.sum(axis=1)
    pivot.index = [f"{m}월" for m in pivot.index]
    total = pivot.sum(axis=0)
    total.name = "합계"
    pivot = pd.concat([pivot, total.to_frame().T])
    return pivot


# ---------------- 유찰경쟁강도 ----------------

def _fail_rate_table(notice: pd.DataFrame, group_col: str, order=None) -> pd.DataFrame:
    rows = []
    groups = order if order else sorted(notice[group_col].dropna().unique())
    for g in groups:
        sub = notice[notice[group_col] == g]
        total = len(sub)
        if total == 0:
            rows.append([g, 0, 0, 0.0, 0, 0.0, 0.0])
            continue
        fail = (sub["개찰상태"].isin(["유찰", "취소"])).sum()
        undisclosed = (sub["개찰상태"] == "결과미공개").sum()
        avg_bidders = sub["참여업체수"].mean()
        rows.append([g, total, fail, fail / total, undisclosed, undisclosed / total, avg_bidders])
    out = pd.DataFrame(rows, columns=["구분", "전체공고수", "유찰·취소", "유찰·취소율", "결과미공개", "결과미공개율", "평균참여업체수"])
    tot_row = pd.DataFrame([[
        "합계", out["전체공고수"].sum(), out["유찰·취소"].sum(),
        out["유찰·취소"].sum() / out["전체공고수"].sum() if out["전체공고수"].sum() else 0,
        out["결과미공개"].sum(),
        out["결과미공개"].sum() / out["전체공고수"].sum() if out["전체공고수"].sum() else 0,
        (out["전체공고수"] * out["평균참여업체수"]).sum() / out["전체공고수"].sum() if out["전체공고수"].sum() else 0,
    ]], columns=out.columns)
    return pd.concat([out, tot_row], ignore_index=True)


def build_fail_rate_analysis(df: pd.DataFrame) -> dict:
    notice = notice_level_df(df)
    type_order = ["재건축", "재개발", "재개발(촉진)", "재개발(생활권)", "도시환경정비",
                  "가로주택정비", "소규모재건축", "소규모재개발", "자율주택정비", "기타"]
    result = {
        "①사업유형별": _fail_rate_table(notice, "사업유형", type_order),
        "②구분별": _fail_rate_table(notice, "구분", CATEGORY_ORDER),
        "⑤발주주체별": _fail_rate_table(notice, "발주주체", ["조합", "추진위원회", "신탁사"]),
    }

    # ③ 시도별 (상위 지역, 조합/추진위/신탁사 발주만 - 원본과 동일하게 신탁사 등 특정 발주주체만 포함된 것으로 보이므로
    #    실제로는 전체 시도 중 상위 N개만 노출. 여기서는 공고수 내림차순 상위 8개만.
    sido_order = notice["시도"].value_counts().index.tolist()
    result["③시도별"] = _fail_rate_table(notice, "시도", sido_order)

    # ④ 참여업체수 분포 (정상 개찰건만)
    normal = notice[notice["개찰상태"] == "정상"]
    bins = [0, 1, 3, 6, 10, np.inf]
    labels = ["1개(단독)", "2~3개", "4~6개", "7~10개", "11개 이상"]
    dist = pd.cut(normal["참여업체수"], bins=bins, labels=labels).value_counts().reindex(labels, fill_value=0)
    dist_df = dist.reset_index()
    dist_df.columns = ["참여업체수 구간", "공고수"]
    dist_df["비중"] = dist_df["공고수"] / dist_df["공고수"].sum()
    result["④참여업체수분포"] = dist_df

    # ⑥ 발주주체 x 사업유형 결과미공개율 교차표
    proj_group = notice.copy()
    proj_group["유형그룹"] = proj_group["사업유형"].map(lambda x: PROJECT_TYPE_GROUPS.get(x, "기타"))
    cross = proj_group.groupby(["발주주체", "유형그룹"]).apply(
        lambda g: (g["개찰상태"] == "결과미공개").sum() / len(g) if len(g) else np.nan
    ).unstack()
    cols = [c for c in ["재건축", "재개발계", "가로주택정비", "소규모정비계"] if c in cross.columns]
    result["⑥교차표"] = cross.reindex(index=["조합", "추진위원회", "신탁사"])[cols]

    # ⑦ 가로주택정비 x 조합, 발주기관별 결과미공개 상세 (2건 이상 발주)
    gj = notice[(proj_group["유형그룹"] == "가로주택정비") & (notice["발주주체"] == "조합")]
    g7 = gj.groupby("발주기관").agg(
        전체공고수=("개찰상태", "count"),
        비공개건수=("개찰상태", lambda s: (s == "결과미공개").sum()),
    )
    g7 = g7[g7["전체공고수"] >= 2].copy()
    g7["비공개율"] = g7["비공개건수"] / g7["전체공고수"]
    g7 = g7.sort_values("비공개율", ascending=False)
    tot = pd.DataFrame([[g7["전체공고수"].sum(), g7["비공개건수"].sum(),
                          g7["비공개건수"].sum() / g7["전체공고수"].sum() if g7["전체공고수"].sum() else 0]],
                        columns=["전체공고수", "비공개건수", "비공개율"], index=["합계"])
    result["⑦가로주택조합상세"] = pd.concat([g7, tot])

    return result


# ---------------- 발주주체지역분포 ----------------

def build_orderer_region(df: pd.DataFrame) -> dict:
    proj = project_level_df(df)

    sido_pv = _pivot_by(proj, "시도")
    tot = sido_pv.sum(numeric_only=True)
    tot["신탁사비중"] = tot["신탁사"] / tot["사업장수"] if tot["사업장수"] else 0
    tot.name = "합계"
    sido_pv = pd.concat([sido_pv, tot.to_frame().T])

    seoul = proj[proj["시도"] == "서울시"].copy()
    seoul["자치구"] = seoul["지역"].map(lambda r: (re.search(r"[가-힣]+구", r).group(0) if re.search(r"[가-힣]+구", r) else ""))
    gu_pv = _pivot_by(seoul, "자치구") if len(seoul) else pd.DataFrame()

    return {"①시도별": sido_pv, "②서울자치구별": gu_pv}


def _pivot_by(sub: pd.DataFrame, group_col: str) -> pd.DataFrame:
    sub = sub.copy()
    # "지역주택조합"은 별도 칸이 없는 표이므로 "기타"로 접어서 집계
    sub["발주주체_표시"] = sub["발주주체"].replace({"지역주택조합": "기타"})
    pv = sub.pivot_table(index=group_col, columns="발주주체_표시", values="사업키(발주기관)", aggfunc="count", fill_value=0)
    for c in ["조합", "추진위원회", "신탁사", "기타"]:
        if c not in pv.columns:
            pv[c] = 0
    pv = pv[["조합", "추진위원회", "신탁사", "기타"]]
    pv["사업장수"] = pv.sum(axis=1)
    pv["신탁사비중"] = pv["신탁사"] / pv["사업장수"]
    return pv.sort_values("사업장수", ascending=False)


# ---------------- 신탁사별사업장 ----------------

def build_trust_projects(df: pd.DataFrame) -> pd.DataFrame:
    proj = project_level_df(df)
    trust = proj[proj["발주주체"] == "신탁사"].copy()
    trust["사업장명(추출)"] = trust["사업키(발주기관)"].map(lambda k: k.split("|")[-1] if "|" in k else "")
    out = trust[["발주기관", "지역", "사업장명(추출)"]].drop_duplicates(subset=["발주기관", "지역", "사업장명(추출)"])
    out = out.sort_values("발주기관")
    counts = trust.groupby("발주기관")["사업키(발주기관)"].nunique().sort_values(ascending=False)
    counts_df = counts.reset_index()
    counts_df.columns = ["신탁사", "사업장수"]
    return {"목록": out, "신탁사별건수": counts_df}


# ---------------- 가격분석 ----------------

def build_price_analysis(df: pd.DataFrame) -> dict:
    unit = df[(df["금액유형"] == "단가") & (df["세부내역"] == "설계자")].copy()
    unit["입찰금액(원)"] = pd.to_numeric(unit["입찰금액(원)"], errors="coerce")
    unit = unit.dropna(subset=["입찰금액(원)"])

    def _agg(sub, group_col):
        g = sub.groupby(group_col)["입찰금액(원)"].agg(응찰수="count", 평균단가="mean", 최저="min", 최고="max")
        return g.sort_values("응찰수", ascending=False)

    by_sido = _agg(unit, "시도")
    by_type = _agg(unit, "사업유형")

    total = df.copy()
    total["입찰금액(원)"] = pd.to_numeric(total["입찰금액(원)"], errors="coerce")
    total_amt = total[(total["금액유형"] == "총액")].dropna(subset=["입찰금액(원)"])
    by_category = total_amt.groupby("구분")["입찰금액(원)"].agg(응찰수="count", 평균총액="mean", 최저="min", 최고="max")
    by_category = by_category.sort_values("평균총액", ascending=False)

    return {"①단가_시도별": by_sido, "②단가_사업유형별": by_type, "③총액_구분별": by_category}


# ---------------- 파이프라인 ----------------

def build_pipeline(df: pd.DataFrame) -> dict:
    proj = project_level_df(df)
    notice = notice_level_df(df)

    def stage(key):
        cats = set(notice[notice["사업키(발주기관)"] == key]["구분"])
        if cats and cats <= EARLY_ONLY_CATEGORIES:
            return "설계·엔지니어링만 (후속 대기)"
        if cats & LATE_STAGE_BROAD:
            return "후반 단계 진입 (감리·토목·PM 등)"
        return "기타 (전반 단계 공고 없음)"

    proj = proj.copy()
    proj["상태"] = proj["사업키(발주기관)"].map(stage)
    stage_counts = proj["상태"].value_counts().reindex(
        ["설계·엔지니어링만 (후속 대기)", "후반 단계 진입 (감리·토목·PM 등)", "기타 (전반 단계 공고 없음)"], fill_value=0
    )

    waiting_keys = proj[proj["상태"] == "설계·엔지니어링만 (후속 대기)"]["사업키(발주기관)"]
    cand_rows = []
    for key in waiting_keys:
        sub = notice[notice["사업키(발주기관)"] == key].sort_values("공고일시")
        if sub.empty:
            continue
        first = sub.iloc[0]
        last = sub.iloc[-1]
        cats = ", ".join(sorted(set(sub["구분"])))
        cand_rows.append([
            first["지역"], first["발주기관"], first["발주주체"],
            first["공고일시"].strftime("%m-%d") if pd.notna(first["공고일시"]) else "",
            last["공고일시"].strftime("%m-%d") if pd.notna(last["공고일시"]) else "",
            cats, len(sub),
        ])
    cand_df = pd.DataFrame(cand_rows, columns=["지역", "발주기관", "발주주체", "첫공고", "최근공고", "발주한 구분", "공고수"])
    cand_df = cand_df.sort_values("최근공고", ascending=False).head(50)

    return {"①진행단계": stage_counts, "②후속임박후보": cand_df}


# ---------------- 구분별시기패턴 ----------------

def build_category_timing(df: pd.DataFrame) -> pd.DataFrame:
    notice = notice_level_df(df)
    multi_keys = notice.groupby("사업키(발주기관)")["공고일시"].transform("count")
    multi = notice[multi_keys >= 2].copy()

    rows_out = []
    for key, sub in multi.groupby("사업키(발주기관)"):
        sub = sub.sort_values("공고일시").reset_index(drop=True)
        n = len(sub)
        first_date = sub.iloc[0]["공고일시"]
        for order_idx, (_, row) in enumerate(sub.iterrows()):
            rel_order = (order_idx + 1) / n
            elapsed = (row["공고일시"] - first_date).days if pd.notna(row["공고일시"]) else 0
            rows_out.append({
                "구분": row["구분"], "상대순서": rel_order, "경과일": elapsed,
                "is_first": order_idx == 0, "is_last": order_idx == n - 1,
            })
    timing_df = pd.DataFrame(rows_out)
    if timing_df.empty:
        return pd.DataFrame()

    agg = timing_df.groupby("구분").agg(
        표본수=("상대순서", "count"),
        평균상대순서=("상대순서", "mean"),
        평균경과일=("경과일", "mean"),
        첫공고등장횟수=("is_first", "sum"),
        마지막공고등장횟수=("is_last", "sum"),
    ).sort_values("평균상대순서")
    n_projects = multi["사업키(발주기관)"].nunique()
    return agg, n_projects


# ---------------- 업체네트워크 ----------------

def build_company_network(df: pd.DataFrame) -> dict:
    notice_bidders = df[df["업체명"].notna() & (df["업체명"].astype(str).str.strip() != "")].copy()

    top_bidders = notice_bidders["업체명"].value_counts().head(20)
    main_field = []
    for company in top_bidders.index:
        sub = notice_bidders[notice_bidders["업체명"] == company]
        cat = sub["구분"].mode().iloc[0] if not sub["구분"].mode().empty else ""
        det = sub["세부내역"].mode().iloc[0] if not sub["세부내역"].mode().empty else ""
        main_field.append(f"{cat}/{det}" if det else cat)
    top_bidders_df = pd.DataFrame({"업체명": top_bidders.index, "응찰횟수": top_bidders.values, "주력분야": main_field})

    pair_counts = {}
    for notice_id, sub in notice_bidders.groupby("공고ID"):
        companies = sorted(sub["업체명"].unique())
        for i in range(len(companies)):
            for j in range(i + 1, len(companies)):
                key = (companies[i], companies[j])
                pair_counts[key] = pair_counts.get(key, 0) + 1
    pairs = sorted(pair_counts.items(), key=lambda x: -x[1])[:20]
    pairs_df = pd.DataFrame([(a, b, c) for (a, b), c in pairs], columns=["업체A", "업체B", "동시등장"])

    return {"②응찰상위업체": top_bidders_df, "③동시등장상위쌍": pairs_df}




# ============================================================
# --- write_excel.py ---
# ============================================================

TITLE_FONT = Font(name="맑은 고딕", bold=True, size=13, color="1F3864")
NOTE_FONT = Font(name="맑은 고딕", size=9, color="666666", italic=True)
SUBTITLE_FONT = Font(name="맑은 고딕", bold=True, size=11, color="1F3864")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
BODY_FONT = Font(name="맑은 고딕", size=10)


def _write_title(ws, row, text, font=TITLE_FONT):
    ws.cell(row=row, column=2, value=text).font = font
    return row + 2


def _write_note(ws, row, text):
    ws.cell(row=row, column=2, value=text).font = NOTE_FONT
    return row + 2


def _write_df(ws, row, df: pd.DataFrame, start_col=2, index_label=None, pct_cols=None, round_cols=None):
    """DataFrame을 (row, start_col) 위치부터 헤더+데이터로 기록. 인덱스가 있으면 첫 컬럼으로 포함."""
    pct_cols = pct_cols or []
    round_cols = round_cols or []
    has_index_name = df.index.name is not None or index_label is not None
    cols = ([index_label or df.index.name or ""] if has_index_name else []) + list(df.columns)

    for j, col_name in enumerate(cols):
        c = ws.cell(row=row, column=start_col + j, value=col_name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1

    for idx, data_row in df.iterrows():
        j = 0
        if has_index_name:
            ws.cell(row=row, column=start_col + j, value=idx).font = BODY_FONT
            j += 1
        for col_name, val in data_row.items():
            if col_name in pct_cols and pd.notna(val):
                val = round(float(val), 4)
            elif col_name in round_cols and pd.notna(val):
                val = round(float(val), 1)
            cell = ws.cell(row=row, column=start_col + j, value=val)
            cell.font = BODY_FONT
            if col_name in pct_cols:
                cell.number_format = "0.0%"
            elif isinstance(val, float):
                cell.number_format = "#,##0"
            j += 1
        row += 1
    return row + 1


def write_summary_sheet(ws, summary: dict):
    row = 2
    row = _write_title(ws, row, "요약 대시보드 · 데이터가 바뀌면 자동 갱신")
    row = _write_note(ws, row, '※ 아래 모든 집계는 "공고 단위"(같은 공고 내 여러 응찰업체는 1건으로 처리) 기준입니다.')

    cat_df = summary["구분별"].reset_index()
    cat_df.columns = ["구분", "건수"]
    cat_df.loc[len(cat_df)] = ["합계", summary["총공고수"]]

    type_df = summary["사업유형별"].reset_index()
    type_df.columns = ["사업유형", "건수"]
    type_df.loc[len(type_df)] = ["합계", summary["총공고수"]]

    status_df = summary["개찰상태별"].reset_index()
    status_df.columns = ["개찰상태", "건수"]
    status_df.loc[len(status_df)] = ["합계", summary["총공고수"]]

    _write_df(ws, row, cat_df, start_col=2)
    _write_df(ws, row, type_df, start_col=5)
    _write_df(ws, row, status_df, start_col=8)

    for col in range(2, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14


def write_region_type_trend_sheet(ws, trend: pd.DataFrame):
    row = 2
    row = _write_title(ws, row, "지역 · 유형별 월별 추이")
    row = _write_note(ws, row,
        "사업유형은 재개발 계열(재개발/촉진/생활권/도시환경), 소규모정비 계열(소규모재건축/재개발/자율주택)로 묶었습니다. "
        '"지역유형추이"와 "발주주체지역분포"는 사업장(아파트/조합) 단위로 집계합니다.')
    row = _write_title(ws, row, "① 사업유형별 월별 건수", SUBTITLE_FONT)
    trend_out = trend.reset_index()
    trend_out.columns = ["월"] + list(trend.columns)
    _write_df(ws, row, trend_out, start_col=2)
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 13


def write_fail_rate_sheet(ws, fr: dict):
    row = 2
    row = _write_title(ws, row, "유찰률 · 경쟁강도 분석")
    row = _write_note(ws, row,
        "유찰·취소율 = 실제로 무산된 공고 비율. 결과미공개율 = 개찰은 했으나 결과를 공개하지 않은 공고 비율(실패 아님). "
        "평균참여업체수 = 공고당 응찰업체 수 평균(경쟁강도).")

    row = _write_title(ws, row, "① 사업유형별", SUBTITLE_FONT)
    row = _write_df(ws, row, fr["①사업유형별"].set_index("구분"), pct_cols=["유찰·취소율", "결과미공개율"], round_cols=["평균참여업체수"])

    row = _write_title(ws, row, "② 구분별", SUBTITLE_FONT)
    row = _write_df(ws, row, fr["②구분별"].set_index("구분"), pct_cols=["유찰·취소율", "결과미공개율"], round_cols=["평균참여업체수"])

    row = _write_title(ws, row, "③ 시도별 (상위 지역)", SUBTITLE_FONT)
    row = _write_df(ws, row, fr["③시도별"].head(8).set_index("구분"), pct_cols=["유찰·취소율", "결과미공개율"], round_cols=["평균참여업체수"])

    row = _write_title(ws, row, "④ 공고당 참여업체수 분포 (정상 개찰건만)", SUBTITLE_FONT)
    row = _write_df(ws, row, fr["④참여업체수분포"].set_index("참여업체수 구간"), pct_cols=["비중"])

    row = _write_title(ws, row, "⑤ 발주주체별", SUBTITLE_FONT)
    row = _write_df(ws, row, fr["⑤발주주체별"].set_index("구분"), pct_cols=["유찰·취소율", "결과미공개율"], round_cols=["평균참여업체수"])

    row = _write_title(ws, row, "⑥ 발주주체 × 사업유형 결과미공개율 (교차표)", SUBTITLE_FONT)
    row = _write_df(ws, row, fr["⑥교차표"], pct_cols=list(fr["⑥교차표"].columns))

    row = _write_title(ws, row, "⑦ 가로주택정비×조합 결과미공개 상세 (2건 이상 발주한 조합)", SUBTITLE_FONT)
    row = _write_df(ws, row, fr["⑦가로주택조합상세"], pct_cols=["비공개율"])

    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions["B"].width = 32


def write_orderer_region_sheet(ws, orl: dict):
    row = 2
    row = _write_title(ws, row, "발주주체별 지역 분포")
    row = _write_note(ws, row, "조합/추진위원회/신탁사가 지역별로 어떻게 다른 비중을 차지하는지. 사업장(아파트/조합) 단위 기준.")

    row = _write_title(ws, row, "① 발주주체 × 시도", SUBTITLE_FONT)
    row = _write_df(ws, row, orl["①시도별"], pct_cols=["신탁사비중"])

    row = _write_title(ws, row, "② 서울 자치구별 발주주체 분포 (상위 15구)", SUBTITLE_FONT)
    row = _write_df(ws, row, orl["②서울자치구별"].head(15), pct_cols=["신탁사비중"])

    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14


def write_trust_projects_sheet(ws, tp: dict):
    row = 2
    row = _write_title(ws, row, "신탁사별 관리 사업장 목록")
    row = _write_note(ws, row, "공고명에서 프로젝트명을 추출해 정리했습니다.")
    row = _write_df(ws, row, tp["목록"].reset_index(drop=True), start_col=2)
    row2 = 6
    row2 = _write_df(ws, row2, tp["신탁사별건수"].set_index("신탁사"), start_col=6)
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 26
    for col in ["F", "G"]:
        ws.column_dimensions[col].width = 22


def write_category_timing_sheet(ws, agg: pd.DataFrame, n_projects: int):
    row = 2
    row = _write_title(ws, row, "구분별 시기 패턴 (사업 단계 힌트)")
    row = _write_note(ws, row,
        f'한 사업장이 서로 다른 구분(설계/엔지니어링/감리 등)의 공고를 여러 번 낸 경우, 그 순서를 분석했습니다. '
        f'대상: 공고를 2개 이상 낸 사업장 {n_projects}개. "평균상대순서"는 0에 가까울수록 초반, 1에 가까울수록 후반.')
    _write_df(ws, row, agg, pct_cols=[], round_cols=["평균상대순서", "평균경과일"])
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 16


def write_price_analysis_sheet(ws, pa: dict):
    row = 2
    row = _write_title(ws, row, "가격 분석 (자릿수 그룹별 분리 비교)")
    row = _write_note(ws, row, "단가 그룹(100만원 미만): ㎡당 단가. 총액 그룹(100만원 이상): 계약 총액.")

    row = _write_title(ws, row, "① 설계자 선정 ㎡당 단가 — 시도별", SUBTITLE_FONT)
    row = _write_df(ws, row, pa["①단가_시도별"], round_cols=["평균단가"])

    row = _write_title(ws, row, "② 설계자 선정 ㎡당 단가 — 사업유형별", SUBTITLE_FONT)
    row = _write_df(ws, row, pa["②단가_사업유형별"], round_cols=["평균단가"])

    row = _write_title(ws, row, "③ 계약 총액 — 구분별", SUBTITLE_FONT)
    row = _write_df(ws, row, pa["③총액_구분별"], round_cols=["평균총액"])

    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 16


def write_pipeline_sheet(ws, pl: dict):
    row = 2
    row = _write_title(ws, row, "사업장 파이프라인 · 신규 진입 분석")
    row = _write_title(ws, row, "① 사업장 진행 단계 현황", SUBTITLE_FONT)
    stage_df = pl["①진행단계"].reset_index()
    stage_df.columns = ["상태", "사업장수"]
    row = _write_df(ws, row, stage_df.set_index("상태"), start_col=2)

    row = _write_title(ws, row, "② 후속 공고 임박 후보 (최근 활동 상위 50곳)", SUBTITLE_FONT)
    row = _write_df(ws, row, pl["②후속임박후보"].reset_index(drop=True), start_col=2)

    ws.column_dimensions["B"].width = 40
    for col in range(3, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16


def write_company_network_sheet(ws, net: dict):
    row = 2
    row = _write_title(ws, row, "공동 응찰 업체 분석")
    row = _write_title(ws, row, "① 응찰횟수 상위 업체 (Top 20)", SUBTITLE_FONT)
    row = _write_df(ws, row, net["②응찰상위업체"].set_index("업체명"), start_col=2)

    row2 = 5
    row2 = _write_title(ws, row2, "② 동시등장(공동응찰) 상위 쌍 (Top 20)", SUBTITLE_FONT, )
    _write_df(ws, row2, net["③동시등장상위쌍"], start_col=6)

    ws.column_dimensions["B"].width = 26
    for col in ["F", "G"]:
        ws.column_dimensions[col].width = 24




# ============================================================
# --- build_report_data.py (build() 함수만, Word 보고서용 데이터 추출) ---
# ============================================================
def df_to_records(df, reset=True):
    d = df.reset_index() if reset and df.index.name else df.copy()
    if not reset and df.index.name is None and not isinstance(df.index, pd.RangeIndex):
        d = df.reset_index()
    return json.loads(d.to_json(orient="records", force_ascii=False))


def build_report_data(input_path_or_bytes) -> dict:
    if isinstance(input_path_or_bytes, bytes):
        input_path_or_bytes = io.BytesIO(input_path_or_bytes)
    df = build_master_df(input_path_or_bytes)
    summary = build_summary(df)
    trend = build_region_type_trend(df)
    fr = build_fail_rate_analysis(df)
    orl = build_orderer_region(df)
    tp = build_trust_projects(df)
    timing_agg, n_projects = build_category_timing(df)
    pa = build_price_analysis(df)
    pl = build_pipeline(df)
    net = build_company_network(df)

    notice = notice_level_df(df)
    proj = project_level_df(df)

    period_start = notice["공고일시"].min()
    period_end = notice["공고일시"].max()

    # --- 종합 시사점에 쓸 동적 랭킹들 ---
    type_tbl = fr["①사업유형별"].set_index("구분")
    type_tbl_body = type_tbl.drop(index="합계", errors="ignore")
    top_undisclosed_type = type_tbl_body["결과미공개율"].idxmax()
    top_undisclosed_type_rate = type_tbl_body["결과미공개율"].max()

    orderer_tbl = fr["⑤발주주체별"].set_index("구분").drop(index="합계", errors="ignore")
    most_transparent_orderer = orderer_tbl["결과미공개율"].idxmin()
    least_transparent_orderer = orderer_tbl["결과미공개율"].idxmax()

    overall_fail_rate = float(type_tbl.loc["합계", "유찰·취소율"]) if "합계" in type_tbl.index else None
    overall_undisclosed_rate = float(type_tbl.loc["합계", "결과미공개율"]) if "합계" in type_tbl.index else None

    price_type = pa["②단가_사업유형별"]
    MIN_SAMPLE = 20  # 표본이 너무 적은 유형(예: 재개발(촉진) n=16)은 월별 변동이 커서 헤드라인 통계에서 제외
    price_type_reliable = price_type[price_type["응찰수"] >= MIN_SAMPLE]
    price_type_for_headline = price_type_reliable if not price_type_reliable.empty else price_type
    highest_price_type = price_type_for_headline["평균단가"].idxmax()
    highest_price_type_val = float(price_type_for_headline["평균단가"].max())
    baseline_types = [t for t in ["재건축", "재개발"] if t in price_type.index]
    baseline_avg = float(price_type.loc[baseline_types, "평균단가"].mean()) if baseline_types else None
    price_premium_pct = (
        (highest_price_type_val / baseline_avg - 1) * 100 if baseline_avg else None
    )

    price_sido = pa["①단가_시도별"]
    lowest_sido_price = price_sido["평균단가"].idxmin()
    lowest_sido_price_val = float(price_sido["평균단가"].min())
    highest_sido_price = price_sido["평균단가"].idxmax()
    highest_sido_price_val = float(price_sido["평균단가"].max())

    stage = pl["①진행단계"]
    waiting_count = int(stage.get("설계·엔지니어링만 (후속 대기)", 0))
    waiting_pct = waiting_count / int(stage.sum()) * 100 if stage.sum() else 0

    # 월별 신규 사업장 진입 vs 기존 사업장 후속 공고 분리
    # 신규사업장(월) = 그 달에 "사업장대표행"(사업장의 첫 공고)이 속한 사업장 수
    # 전체공고(월) = 그 달의 전체 공고 건수(공고 단위) / 후속공고 = 전체공고 - 신규사업장
    notice_by_month = notice.groupby(notice["공고일시"].dt.month).size()
    proj_by_month = proj.groupby(proj["공고일시"].dt.month).size()
    monthly_split = pd.DataFrame({"신규사업장": proj_by_month, "전체공고": notice_by_month}).fillna(0)
    monthly_split["후속공고"] = monthly_split["전체공고"] - monthly_split["신규사업장"]
    monthly_split["신규비중"] = monthly_split["신규사업장"] / monthly_split["전체공고"]
    monthly_split = monthly_split.reindex(range(1, 13)).dropna(how="all")
    monthly_split.index = [f"{int(m)}월" for m in monthly_split.index]
    tot = monthly_split.sum(numeric_only=True)
    tot["신규비중"] = tot["신규사업장"] / tot["전체공고"] if tot["전체공고"] else 0
    tot.name = "합계"
    monthly_split = pd.concat([monthly_split, tot.to_frame().T])

    data = {
        "generated_at": datetime.now().strftime("%Y-%m-%d"),
        "period_start": period_start.strftime("%Y-%m-%d") if pd.notna(period_start) else "",
        "period_end": period_end.strftime("%Y-%m-%d") if pd.notna(period_end) else "",
        "총공고수": summary["총공고수"],
        "총사업장수": int(proj.shape[0]),
        "총응찰행수": int(len(df)),

        "구분별": df_to_records(summary["구분별"].reset_index().rename(columns={"index": "구분", "count": "건수", 0: "건수"})),
        "사업유형별": df_to_records(summary["사업유형별"].reset_index().rename(columns={"index": "사업유형", "count": "건수", 0: "건수"})),
        "개찰상태별": df_to_records(summary["개찰상태별"].reset_index().rename(columns={"index": "개찰상태", "count": "건수", 0: "건수"})),

        "지역유형추이": df_to_records(trend.reset_index().rename(columns={"index": "월"})),

        "유찰_사업유형별": df_to_records(fr["①사업유형별"]),
        "유찰_구분별": df_to_records(fr["②구분별"]),
        "유찰_시도별": df_to_records(fr["③시도별"].head(8)),
        "유찰_발주주체별": df_to_records(fr["⑤발주주체별"]),
        "유찰_교차표": df_to_records(fr["⑥교차표"].reset_index()),
        "유찰_가로주택조합상세": df_to_records(fr["⑦가로주택조합상세"].reset_index().rename(columns={"index": "발주기관"}).head(10)),

        "발주주체_시도별": df_to_records(orl["①시도별"].reset_index().rename(columns={"index": "시도"})),
        "발주주체_자치구별": df_to_records(orl["②서울자치구별"].reset_index().rename(columns={"index": "자치구"}).head(10)),

        "신탁사별건수": df_to_records(tp["신탁사별건수"]),

        "가격_단가_시도별": df_to_records(pa["①단가_시도별"].reset_index().rename(columns={"index": "시도"})),
        "가격_단가_유형별": df_to_records(pa["②단가_사업유형별"].reset_index().rename(columns={"index": "사업유형"})),
        "가격_총액_구분별": df_to_records(pa["③총액_구분별"].reset_index().rename(columns={"index": "구분"})),

        "파이프라인_진행단계": df_to_records(stage.reset_index().rename(columns={"index": "상태", "count": "사업장수", 0: "사업장수"})),
        "파이프라인_후속임박": df_to_records(pl["②후속임박후보"].head(20)),
        "파이프라인_월별신규후속": df_to_records(monthly_split.reset_index().rename(columns={"index": "월"})),

        "인사이트": {
            "결과미공개율최고유형": top_undisclosed_type,
            "결과미공개율최고유형_비율": round(float(top_undisclosed_type_rate) * 100, 1),
            "가장투명한발주주체": most_transparent_orderer,
            "가장불투명한발주주체": least_transparent_orderer,
            "전체유찰취소율": round(overall_fail_rate * 100, 1) if overall_fail_rate is not None else None,
            "전체결과미공개율": round(overall_undisclosed_rate * 100, 1) if overall_undisclosed_rate is not None else None,
            "최고단가사업유형": highest_price_type,
            "최고단가사업유형_값": round(highest_price_type_val),
            "최고단가_기준대비프리미엄pct": round(price_premium_pct, 1) if price_premium_pct is not None else None,
            "최저단가지역": lowest_sido_price,
            "최저단가지역_값": round(lowest_sido_price_val),
            "최고단가지역": highest_sido_price,
            "최고단가지역_값": round(highest_sido_price_val),
            "후속대기사업장수": waiting_count,
            "후속대기비율": round(waiting_pct, 1),
        },
    }
    return data


# ============================================================
# --- refresh_report.py 로직 (BytesIO 기반, 웹앱용) ---
# ============================================================

MASTER_SHEET = "2026 정비사업 입찰공고 정리내역"
SHEET_ORDER = [
    MASTER_SHEET, "요약", "기준정보", "지역유형추이", "유찰경쟁강도",
    "발주주체지역분포", "신탁사별사업장", "업체네트워크", "구분별시기패턴",
    "데이터품질", "가격분석", "파이프라인",
]


def refresh_excel_bytes(input_bytes: bytes, progress_cb=None) -> bytes:
    """마스터 xlsx(bytes)를 받아, 11개 집계시트를 재계산한 새 워크북을 bytes로 반환.
    마스터 원본 데이터·기준정보·데이터품질 시트는 그대로 보존됩니다."""

    def _log(msg):
        if progress_cb:
            progress_cb(msg)

    _log("마스터 데이터를 불러오는 중...")
    df = build_master_df(io.BytesIO(input_bytes), MASTER_SHEET)
    _log(f"{len(df)}행(응찰업체별) 로드 완료. 11개 집계시트 계산 중...")

    summary = build_summary(df)
    trend = build_region_type_trend(df)
    fr = build_fail_rate_analysis(df)
    orl = build_orderer_region(df)
    tp = build_trust_projects(df)
    timing_agg, n_projects = build_category_timing(df)
    pa = build_price_analysis(df)
    pl = build_pipeline(df)
    net = build_company_network(df)

    _log("집계시트를 새 워크북에 기록하는 중...")
    wb = openpyxl.load_workbook(io.BytesIO(input_bytes))

    sheet_writers = {
        "요약": lambda ws: write_summary_sheet(ws, summary),
        "지역유형추이": lambda ws: write_region_type_trend_sheet(ws, trend),
        "유찰경쟁강도": lambda ws: write_fail_rate_sheet(ws, fr),
        "발주주체지역분포": lambda ws: write_orderer_region_sheet(ws, orl),
        "신탁사별사업장": lambda ws: write_trust_projects_sheet(ws, tp),
        "구분별시기패턴": lambda ws: write_category_timing_sheet(ws, timing_agg, n_projects),
        "가격분석": lambda ws: write_price_analysis_sheet(ws, pa),
        "파이프라인": lambda ws: write_pipeline_sheet(ws, pl),
        "업체네트워크": lambda ws: write_company_network_sheet(ws, net),
    }

    for sheet_name, writer_fn in sheet_writers.items():
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)
        writer_fn(ws)

    for i, name in enumerate(SHEET_ORDER):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    _log("완료")
    out_buf = io.BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue()
